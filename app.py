from flask import Flask, render_template, request, redirect, session, send_file
from database.models import criar_tabelas
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet
from openpyxl import Workbook
from database.conexao import conectar
from datetime import datetime
from zoneinfo import ZoneInfo
from werkzeug.utils import secure_filename
from zipfile import ZipFile
import os
import re

app = Flask(__name__)

app.secret_key = "ponto_web_2026"

criar_tabelas()

@app.route('/')
def inicio():
    session.clear()
    return render_template('index.html')


# =========================
# ADMINISTRADOR
# =========================

@app.route('/login-administrador', methods=['GET', 'POST'])
def login_administrador():

    if request.method == 'POST':

        login = request.form.get('login')
        senha = request.form.get('senha')

        if login == 'gestor1' and senha == 'gestor2026':

            session.clear()
            session['administrador_id'] = 1
            session['nome_admin'] = 'Administrador'

            return redirect('/dashboard-administrador')

        return render_template(
            'administrador/login_administrador.html',
            erro='Login ou senha inválidos'
        )

    return render_template(
        'administrador/login_administrador.html'
    )


# =========================
# GESTOR DO NÚCLEO
# =========================

@app.route('/login-gestor-nucleo', methods=['GET', 'POST'])
def login_gestor_nucleo():

    if request.method == 'POST':

        login = request.form.get('login')
        senha = request.form.get('senha')

        conn = conectar()
        cursor = conn.cursor()

        cursor.execute("""
            SELECT *
            FROM gestores
            WHERE login = ?
            AND senha = ?
            AND status = 'Ativo'
        """, (
            login,
            senha
        ))

        gestor = cursor.fetchone()

        conn.close()

        if gestor:

            session.clear()
            session['gestor_id'] = gestor[0]
            session['nome_gestor'] = gestor[1]
            session['nucleo_gestor'] = gestor[2]

            # Primeiro acesso
            if gestor[6] == "Novocolab123":
                return redirect('/primeiro-acesso-gestor')

            return redirect('/dashboard-gestor-nucleo')

        return render_template(
            'gestor_nucleo/login_gestor_nucleo.html',
            erro='Login ou senha inválidos'
        )

    return render_template(
        'gestor_nucleo/login_gestor_nucleo.html'
    )


@app.route('/dashboard-gestor-nucleo')
def dashboard_gestor_nucleo():

    if 'gestor_id' not in session:
        return redirect('/login-gestor-nucleo')

    conn = conectar()
    cursor = conn.cursor()

    # Total de colaboradores
    cursor.execute("""
        SELECT COUNT(*)
        FROM colaboradores
        WHERE nucleo = ?
    """, (
        session['nucleo_gestor'],
    ))
    total_colaboradores = cursor.fetchone()[0]

    # Colaboradores ativos
    cursor.execute("""
        SELECT COUNT(*)
        FROM colaboradores
        WHERE nucleo = ?
        AND status = 'Ativo'
    """, (
        session['nucleo_gestor'],
    ))
    colaboradores_ativos = cursor.fetchone()[0]

    # Colaboradores inativos
    cursor.execute("""
        SELECT COUNT(*)
        FROM colaboradores
        WHERE nucleo = ?
        AND status = 'Inativo'
    """, (
        session['nucleo_gestor'],
    ))
    colaboradores_inativos = cursor.fetchone()[0]

    # Ajustes pendentes
    cursor.execute("""
        SELECT COUNT(*)
        FROM ajustes_ponto
        INNER JOIN colaboradores
        ON ajustes_ponto.colaborador_id = colaboradores.id
        WHERE colaboradores.nucleo = ?
        AND ajustes_ponto.status = 'Pendente'
    """, (
        session['nucleo_gestor'],
    ))
    ajustes_pendentes = cursor.fetchone()[0]

    conn.close()

    return render_template(
        'gestor_nucleo/dashboard_gestor_nucleo.html',
        nome=session['nome_gestor'],
        nucleo=session['nucleo_gestor'],
        total_colaboradores=total_colaboradores,
        colaboradores_ativos=colaboradores_ativos,
        colaboradores_inativos=colaboradores_inativos,
        ajustes_pendentes=ajustes_pendentes
    )


# =========================
# CADASTRO DE GESTOR
# =========================

@app.route('/cadastro-gestor')
def cadastro_gestor():

    if 'administrador_id' not in session:
        return redirect('/login-administrador')

    return render_template(
        'administrador/cadastro_gestor.html'
    )


@app.route('/salvar-gestor', methods=['POST'])
def salvar_gestor():

    if 'administrador_id' not in session:
        return redirect('/login-administrador')

    conn = conectar()
    cursor = conn.cursor()

    nome = request.form.get('nome', '')
    nucleo = request.form.get('nucleo', '')
    unidade_exercicio = request.form.get('unidade_exercicio', '')
    celular = request.form.get('celular', '')
    login = request.form.get('login', '')

    try:
        cursor.execute("""
            INSERT INTO gestores (
                nome,
                nucleo,
                unidade_exercicio,
                celular,
                login,
                senha
            )
            VALUES (?, ?, ?, ?, ?, ?)
        """, (
            nome,
            nucleo,
            unidade_exercicio,
            celular,
            login,
            "Novocolab123"
        ))

        conn.commit()
        return redirect('/gestores-cadastrados')
    except Exception as e:
        conn.rollback()
        error_message = 'Erro ao cadastrar gestor. Verifique se o login já existe.'
        if 'UNIQUE constraint failed' in str(e):
            error_message = 'Já existe um gestor cadastrado com esse login.'

        return render_template(
            'administrador/cadastro_gestor.html',
            erro=error_message,
            nome=nome,
            nucleo=nucleo,
            unidade_exercicio=unidade_exercicio,
            celular=celular,
            login=login
        )
    finally:
        conn.close()

# =========================
# GESTORES CADASTRADOS
# =========================

@app.route('/editar-gestor/<int:id>')
def editar_gestor(id):

    if 'administrador_id' not in session:
        return redirect('/login-administrador')

    conn = conectar()
    cursor = conn.cursor()

    cursor.execute("""
        SELECT *
        FROM gestores
        WHERE id = ?
    """, (
        id,
    ))

    gestor = cursor.fetchone()

    conn.close()

    return render_template(
        'administrador/editar_gestor.html',
        gestor=gestor
    )


@app.route('/atualizar-gestor/<int:id>', methods=['POST'])
def atualizar_gestor(id):

    if 'administrador_id' not in session:
        return redirect('/login-administrador')

    conn = conectar()
    cursor = conn.cursor()

    cursor.execute("""
        UPDATE gestores
        SET
            nome = ?,
            nucleo = ?,
            login = ?
        WHERE id = ?
    """, (
        request.form['nome'],
        request.form['nucleo'],
        request.form['login'],
        id
    ))

    conn.commit()
    conn.close()

    return redirect('/gestores-cadastrados')


@app.route('/cancelar-gestor/<int:id>', methods=['POST'])
def cancelar_gestor(id):

    if 'administrador_id' not in session:
        return redirect('/login-administrador')

    motivo = request.form.get('motivo_cancelamento', '').strip()

    conn = conectar()
    cursor = conn.cursor()

    cursor.execute("""
        UPDATE gestores
        SET status = 'Inativo', cancel_observacao = ?
        WHERE id = ?
    """, (
        motivo,
        id,
    ))

    conn.commit()
    conn.close()

    return redirect('/gestores-cadastrados')

@app.route('/recadastrar-gestor/<int:id>', methods=['POST'])
def recadastrar_gestor(id):

    if 'administrador_id' not in session:
        return redirect('/login-administrador')

    conn = conectar()
    cursor = conn.cursor()

    cursor.execute("""
        UPDATE gestores
        SET status = 'Ativo', cancel_observacao = NULL
        WHERE id = ?
    """, (
        id,
    ))

    conn.commit()
    conn.close()

    return redirect('/gestores-cadastrados')

@app.route('/resetar-senha-gestor/<int:id>')
def resetar_senha_gestor(id):

    if 'administrador_id' not in session:
        return redirect('/login-administrador')

    conn = conectar()
    cursor = conn.cursor()

    cursor.execute("""
        UPDATE gestores
        SET senha = ?
        WHERE id = ?
    """, (
        "Novocolab123",
        id
    ))

    conn.commit()
    conn.close()

    return redirect('/gestores-cadastrados')

@app.route('/resetar-senha-colaborador/<int:id>')
def resetar_senha_colaborador(id):

    if 'administrador_id' not in session:
        return redirect('/login-administrador')

    conn = conectar()
    cursor = conn.cursor()

    cursor.execute("""
        UPDATE colaboradores
        SET senha = ?
        WHERE id = ?
    """, (
        "Novocolab123",
        id
    ))

    conn.commit()
    conn.close()

    return redirect('/colaboradores-cadastrados')

@app.route('/gestores-cadastrados')
def gestores_cadastrados():

    if 'administrador_id' not in session:
        return redirect('/login-administrador')

    conn = conectar()
    cursor = conn.cursor()

    cursor.execute("""
        SELECT *
        FROM gestores
        ORDER BY nome
    """)

    gestores = cursor.fetchall()

    conn.close()

    return render_template(
        'administrador/gestores_cadastrados.html',
        gestores=gestores
    )


# =========================
# COLABORADOR
# =========================

@app.route('/login-colaborador', methods=['GET', 'POST'])
def login_colaborador():

    if request.method == 'POST':

        login = request.form.get('login')
        senha = request.form.get('senha')

        conn = conectar()
        cursor = conn.cursor()

        cursor.execute("""
            SELECT *
            FROM colaboradores
            WHERE login = ?
            AND senha = ?
            AND status = 'Ativo'
        """, (
            login,
            senha
        ))

        colaborador = cursor.fetchone()

        conn.close()

        if colaborador:

            session.clear()
            session['colaborador_id'] = colaborador[0]
            session['nome'] = colaborador[1]

            if colaborador[11] == "Novocolab123":
                return redirect('/primeiro-acesso')

            return redirect('/ponto')

        return render_template(
            'colaborador/login_colaborador.html',
            erro='Login ou senha inválidos'
        )

    return render_template(
        'colaborador/login_colaborador.html'
    )


# =========================
# CADASTRO DE COLABORADOR
# =========================

@app.route('/cadastro-colaborador')
def cadastro_colaborador():

    if 'administrador_id' in session:
        return render_template(
            'administrador/cadastro_colaborador.html'
        )

    if 'gestor_id' in session:
        return render_template(
            'gestor_nucleo/cadastro_colaborador.html',
            nucleo=session['nucleo_gestor']
        )

    return redirect('/')


# =========================
# DASHBOARD ADMINISTRADOR
# =========================

@app.route('/dashboard-administrador')
def dashboard_administrador():

    conn = conectar()
    cursor = conn.cursor()

    # Total de colaboradores
    cursor.execute("""
        SELECT COUNT(*)
        FROM colaboradores
    """)
    total_colaboradores = cursor.fetchone()[0]

    # Colaboradores ativos
    cursor.execute("""
        SELECT COUNT(*)
        FROM colaboradores
        WHERE status = 'Ativo'
    """)
    colaboradores_ativos = cursor.fetchone()[0]

    # Colaboradores inativos
    cursor.execute("""
        SELECT COUNT(*)
        FROM colaboradores
        WHERE status = 'Inativo'
    """)
    colaboradores_inativos = cursor.fetchone()[0]

    # Total de gestores
    cursor.execute("""
        SELECT COUNT(*)
        FROM gestores
        WHERE status = 'Ativo'
    """)
    total_gestores = cursor.fetchone()[0]

    # Total de núcleos
    cursor.execute("""
        SELECT COUNT(DISTINCT nucleo)
        FROM colaboradores
    """)
    total_nucleos = cursor.fetchone()[0]

    # Ajustes pendentes
    cursor.execute("""
        SELECT COUNT(*)
        FROM ajustes_ponto
        WHERE status = 'Pendente'
    """)
    ajustes_pendentes = cursor.fetchone()[0]

    conn.close()

    return render_template(
        'administrador/dashboard_administrador.html',
        total_colaboradores=total_colaboradores,
        colaboradores_ativos=colaboradores_ativos,
        colaboradores_inativos=colaboradores_inativos,
        total_gestores=total_gestores,
        total_nucleos=total_nucleos,
        ajustes_pendentes=ajustes_pendentes
    )

# =========================
# COLABORADORES CADASTRADOS
# =========================

@app.route('/colaboradores-cadastrados')
def colaboradores_cadastrados():

    conn = conectar()
    cursor = conn.cursor()

    # Administrador vê todos os colaboradores
    if 'administrador_id' in session:
        cursor.execute("""
            SELECT *
            FROM colaboradores
            ORDER BY nome
        """)

    elif 'gestor_id' in session:
        cursor.execute("""
            SELECT *
            FROM colaboradores
            WHERE nucleo = ?
            ORDER BY nome
        """, (
            session['nucleo_gestor'],
        ))

    else:
        return redirect('/login-administrador')

    colaboradores = cursor.fetchall()

    conn.close()

    if 'administrador_id' in session:
        return render_template(
            'administrador/colaboradores.html',
            colaboradores=colaboradores
        )

    return render_template(
        'gestor_nucleo/colaboradores.html',
        colaboradores=colaboradores
    )


# =========================
# PONTO
# =========================

@app.route('/ponto')
def ponto():

    if 'colaborador_id' not in session:
        return redirect('/login-colaborador')

    conn = conectar()
    cursor = conn.cursor()

    data_hoje = datetime.now(
        ZoneInfo("America/Sao_Paulo")
    ).strftime('%d/%m/%Y')

    cursor.execute("""
        SELECT
            id,
            entrada,
            saida_final
        FROM registros_ponto
        WHERE colaborador_id = ?
        AND data = ?
    """, (
        session['colaborador_id'],
        data_hoje
    ))

    registro = cursor.fetchone()

    conn.close()

    entrada = "--:--"
    saida_final = "--:--"
    entrada_registrada = False
    saida_registrada = False

    if registro:
        _, entrada_salva, saida_salva = registro
        entrada = entrada_salva or "--:--"
        saida_final = saida_salva or "--:--"
        entrada_registrada = bool((entrada_salva or '').strip())
        saida_registrada = bool((saida_salva or '').strip())

    return render_template(
        'colaborador/ponto.html',
        nome=session['nome'],
        entrada=entrada,
        saida_final=saida_final,
        entrada_registrada=entrada_registrada,
        saida_registrada=saida_registrada
    )


# =========================
# AJUSTE DE PONTO
# =========================

@app.route('/ajuste-ponto')
def ajuste_ponto():

    return render_template(
        'colaborador/ajuste_ponto.html'
    )


# =========================
# SOLICITAÇÕES DE AJUSTE
# =========================

@app.route('/solicitacoes-ajuste')
def solicitacoes_ajuste():

    conn = conectar()
    cursor = conn.cursor()

    # Administrador vê todos
    if 'gestor_id' not in session:

        cursor.execute("""
            SELECT
                ajustes_ponto.id,
                colaboradores.nome,
                colaboradores.nucleo,
                ajustes_ponto.data,
                ajustes_ponto.motivo,
                ajustes_ponto.status
            FROM ajustes_ponto
            INNER JOIN colaboradores
            ON ajustes_ponto.colaborador_id = colaboradores.id
            ORDER BY ajustes_ponto.id DESC
        """)

    # Gestor do Núcleo vê apenas os ajustes do seu núcleo
    else:

        cursor.execute("""
            SELECT
                ajustes_ponto.id,
                colaboradores.nome,
                colaboradores.nucleo,
                ajustes_ponto.data,
                ajustes_ponto.motivo,
                ajustes_ponto.status
            FROM ajustes_ponto
            INNER JOIN colaboradores
            ON ajustes_ponto.colaborador_id = colaboradores.id
            WHERE colaboradores.nucleo = ?
            ORDER BY ajustes_ponto.id DESC
        """, (
            session['nucleo_gestor'],
        ))

    ajustes = cursor.fetchall()

    conn.close()

    if 'administrador_id' in session:
        template_name = 'administrador/solicitacoes_ajuste.html'
    elif 'gestor_id' in session:
        template_name = 'gestor_nucleo/solicitacoes_ajuste.html'
    else:
        return redirect('/login-gestor-nucleo')

    return render_template(
        template_name,
        ajustes=ajustes
    )

# =========================
# APROVAR AJUSTE
# =========================

@app.route('/aprovar-ajuste/<int:id>')
def aprovar_ajuste(id):

    conn = conectar()
    cursor = conn.cursor()

    # Segurança para Gestor do Núcleo
    if 'gestor_id' in session:

        cursor.execute("""
            SELECT ajustes_ponto.id
            FROM ajustes_ponto
            INNER JOIN colaboradores
            ON ajustes_ponto.colaborador_id = colaboradores.id
            WHERE ajustes_ponto.id = ?
            AND colaboradores.nucleo = ?
        """, (
            id,
            session['nucleo_gestor']
        ))

        permissao = cursor.fetchone()

        if not permissao:
            conn.close()
            return redirect('/solicitacoes-ajuste')

    cursor.execute("""
        UPDATE ajustes_ponto
        SET status = 'Aprovado'
        WHERE id = ?
    """, (id,))

    conn.commit()
    conn.close()

    return redirect('/solicitacoes-ajuste')


# =========================
# REPROVAR AJUSTE
# =========================

@app.route('/reprovar-ajuste/<int:id>')
def reprovar_ajuste(id):

    conn = conectar()
    cursor = conn.cursor()

    # Segurança para Gestor do Núcleo
    if 'gestor_id' in session:

        cursor.execute("""
            SELECT ajustes_ponto.id
            FROM ajustes_ponto
            INNER JOIN colaboradores
            ON ajustes_ponto.colaborador_id = colaboradores.id
            WHERE ajustes_ponto.id = ?
            AND colaboradores.nucleo = ?
        """, (
            id,
            session['nucleo_gestor']
        ))

        permissao = cursor.fetchone()

        if not permissao:
            conn.close()
            return redirect('/solicitacoes-ajuste')

    cursor.execute("""
        UPDATE ajustes_ponto
        SET status = 'Reprovado'
        WHERE id = ?
    """, (id,))

    conn.commit()
    conn.close()

    return redirect('/solicitacoes-ajuste')


# =========================
# MEUS AJUSTES
# =========================

@app.route('/meus-ajustes')
def meus_ajustes():

    if 'colaborador_id' not in session:
        return redirect('/login-colaborador')

    conn = conectar()
    cursor = conn.cursor()

    cursor.execute("""
        SELECT
            data,
            motivo,
            status
        FROM ajustes_ponto
        WHERE colaborador_id = ?
        ORDER BY id DESC
    """, (
        session['colaborador_id'],
    ))

    ajustes = cursor.fetchall()

    conn.close()

    return render_template(
        'colaborador/meus_ajustes.html',
        ajustes=ajustes
    )


# =========================
# CALENDÁRIO
# =========================

@app.route('/calendario')
def calendario():

    conn = conectar()
    cursor = conn.cursor()

    if 'gestor_id' in session:
        cursor.execute("""
            SELECT id, titulo, tipo, data, descricao
            FROM eventos
            WHERE nucleo = ?
            ORDER BY data
        """, (
            session['nucleo_gestor'],
        ))
        eventos = cursor.fetchall()
        conn.close()
        return render_template('gestor_nucleo/calendario.html', eventos=eventos, nucleo=session['nucleo_gestor'])

    if 'administrador_id' in session:
        cursor.execute("""
            SELECT id, titulo, tipo, data, descricao, nucleo
            FROM eventos
            ORDER BY data
        """)
        eventos = cursor.fetchall()
        conn.close()
        return render_template('administrador/calendario.html', eventos=eventos)

    conn.close()
    return redirect('/login-gestor-nucleo')


@app.route('/adicionar-evento', methods=['POST'])
def adicionar_evento():

    if 'gestor_id' not in session:
        return redirect('/login-gestor-nucleo')

    titulo = request.form.get('titulo', '').strip()
    tipo = request.form.get('tipo', '').strip()
    data = request.form.get('data', '').strip()
    descricao = request.form.get('descricao', '').strip()

    if not titulo or not tipo or not data:
        return redirect('/calendario')

    try:
        data_formatada = datetime.strptime(data, '%Y-%m-%d').strftime('%d/%m/%Y')
    except ValueError:
        data_formatada = data

    conn = conectar()
    cursor = conn.cursor()

    cursor.execute("""
        INSERT INTO eventos (
            titulo,
            tipo,
            data,
            descricao,
            nucleo
        ) VALUES (?, ?, ?, ?, ?)
    """, (
        titulo,
        tipo,
        data_formatada,
        descricao,
        session['nucleo_gestor']
    ))

    conn.commit()
    conn.close()

    return redirect('/calendario')


@app.route('/apagar-evento', methods=['POST'])
def apagar_evento():

    if 'gestor_id' not in session:
        return redirect('/login-gestor-nucleo')

    evento_id = request.form.get('event_id')
    if not evento_id:
        return redirect('/calendario')

    conn = conectar()
    cursor = conn.cursor()

    cursor.execute("""
        DELETE FROM eventos
        WHERE id = ?
        AND nucleo = ?
    """, (
        evento_id,
        session['nucleo_gestor']
    ))

    conn.commit()
    conn.close()

    return redirect('/calendario')


def _buscar_status_folhas(cursor, nucleo=None):
    if nucleo is None:
        cursor.execute("""
            SELECT id, nome, folha_assinada_path
            FROM colaboradores
            WHERE status = 'Ativo'
            ORDER BY nome
        """)
    else:
        cursor.execute("""
            SELECT id, nome, folha_assinada_path
            FROM colaboradores
            WHERE status = 'Ativo'
            AND nucleo = ?
            ORDER BY nome
        """, (nucleo,))

    return [
        {
            'id': row[0],
            'nome': row[1],
            'anexado': bool((row[2] or '').strip())
        }
        for row in cursor.fetchall()
    ]


# =========================
# RELATÓRIOS
# =========================

@app.route('/relatorios')
def relatorios():

    conn = conectar()
    cursor = conn.cursor()

    # Administrador vê todos
    if 'gestor_id' not in session:

        cursor.execute("""
            SELECT id, nome
            FROM colaboradores
            WHERE status = 'Ativo'
            ORDER BY nome
        """)

    # Gestor do Núcleo vê apenas o seu núcleo
    else:

        cursor.execute("""
            SELECT id, nome
            FROM colaboradores
            WHERE status = 'Ativo'
            AND nucleo = ?
            ORDER BY nome
        """, (
            session['nucleo_gestor'],
        ))

    colaboradores = cursor.fetchall()
    status_folhas = _buscar_status_folhas(cursor, session.get('nucleo_gestor') if 'gestor_id' in session else None)

    conn.close()

    if 'administrador_id' in session:
        template_name = 'administrador/relatorios.html'
    elif 'gestor_id' in session:
        template_name = 'gestor_nucleo/relatorios.html'
    else:
        return redirect('/login-gestor-nucleo')

    return render_template(
        template_name,
        colaboradores=colaboradores,
        registros=None,
        todos=False,
        status_folhas=status_folhas
    )


# =========================
# VISUALIZAR RELATÓRIO
# =========================

@app.route('/visualizar-relatorio', methods=['POST'])
def visualizar_relatorio():

    colaborador_id = request.form['colaborador_id']

    mes = request.form.get('mes')

    conn = conectar()
    cursor = conn.cursor()

    # Lista de colaboradores
    if 'gestor_id' not in session:

        cursor.execute("""
            SELECT id, nome
            FROM colaboradores
            WHERE status = 'Ativo'
            ORDER BY nome
        """)

    else:

        cursor.execute("""
            SELECT id, nome
            FROM colaboradores
            WHERE status = 'Ativo'
            AND nucleo = ?
            ORDER BY nome
        """, (
            session['nucleo_gestor'],
        ))

    colaboradores = cursor.fetchall()
    status_folhas = _buscar_status_folhas(cursor, session.get('nucleo_gestor') if 'gestor_id' in session else None)

    # Segurança adicional para Gestor do Núcleo
    if 'gestor_id' in session:

        if colaborador_id == 'all':
            conn.close()
            return render_template(
                'gestor_nucleo/relatorios.html',
                colaboradores=colaboradores,
                registros=None,
                erro='Opção inválida.',
                mes=mes,
                todos=False
            )

        cursor.execute("""
            SELECT id
            FROM colaboradores
            WHERE id = ?
            AND nucleo = ?
        """, (
            colaborador_id,
            session['nucleo_gestor']
        ))

        permissao = cursor.fetchone()

        if not permissao:

            conn.close()
            return render_template(
                'gestor_nucleo/relatorios.html',
                colaboradores=colaboradores,
                registros=None,
                erro='Você não possui permissão para visualizar este colaborador.',
                mes=mes,
                todos=False
            )

    # Busca os registros — filtra por mês/ano se fornecido
    if mes:
        # campo <input type="month"> envia 'YYYY-MM'
        try:
            ano, mes_num = mes.split('-')
        except ValueError:
            ano = None
            mes_num = None

    else:
        ano = None
        mes_num = None

    if colaborador_id == 'all':
        if 'administrador_id' not in session:
            conn.close()
            return redirect('/login-administrador')

        if ano and mes_num:
            cursor.execute("""
                SELECT
                    registros_ponto.data,
                    registros_ponto.entrada,
                    registros_ponto.saida_final,
                    COALESCE(NULLIF(registros_ponto.observacao, ''), group_concat(eventos.tipo || ' - ' || eventos.titulo, '; ')) AS observacao,
                    colaboradores.nome
                FROM registros_ponto
                INNER JOIN colaboradores ON colaboradores.id = registros_ponto.colaborador_id
                LEFT JOIN eventos ON eventos.data = registros_ponto.data
                    AND eventos.nucleo = colaboradores.nucleo
                WHERE substr(registros_ponto.data,4,2) = ?
                AND substr(registros_ponto.data,7,4) = ?
                GROUP BY registros_ponto.id
                ORDER BY registros_ponto.id DESC
            """, (
                mes_num,
                ano
            ))
        else:
            cursor.execute("""
                SELECT
                    registros_ponto.data,
                    registros_ponto.entrada,
                    registros_ponto.saida_final,
                    COALESCE(NULLIF(registros_ponto.observacao, ''), group_concat(eventos.tipo || ' - ' || eventos.titulo, '; ')) AS observacao,
                    colaboradores.nome
                FROM registros_ponto
                INNER JOIN colaboradores ON colaboradores.id = registros_ponto.colaborador_id
                LEFT JOIN eventos ON eventos.data = registros_ponto.data
                    AND eventos.nucleo = colaboradores.nucleo
                GROUP BY registros_ponto.id
                ORDER BY registros_ponto.id DESC
            """)

        registros = cursor.fetchall()
        nome_colaborador = 'Todos os colaboradores'
        todos = True

    else:
        if ano and mes_num:
            cursor.execute("""
                SELECT
                    registros_ponto.data,
                    registros_ponto.entrada,
                    registros_ponto.saida_final,
                    COALESCE(NULLIF(registros_ponto.observacao, ''), group_concat(eventos.tipo || ' - ' || eventos.titulo, '; ')) AS observacao
                FROM registros_ponto
                LEFT JOIN eventos ON eventos.data = registros_ponto.data
                    AND eventos.nucleo = (SELECT nucleo FROM colaboradores WHERE id = registros_ponto.colaborador_id)
                WHERE colaborador_id = ?
                AND substr(registros_ponto.data,4,2) = ?
                AND substr(registros_ponto.data,7,4) = ?
                GROUP BY registros_ponto.id
                ORDER BY registros_ponto.id DESC
            """, (
                colaborador_id,
                mes_num,
                ano
            ))

        else:
            cursor.execute("""
                SELECT
                    registros_ponto.data,
                    registros_ponto.entrada,
                    registros_ponto.saida_final,
                    COALESCE(NULLIF(registros_ponto.observacao, ''), group_concat(eventos.tipo || ' - ' || eventos.titulo, '; ')) AS observacao
                FROM registros_ponto
                LEFT JOIN eventos ON eventos.data = registros_ponto.data
                    AND eventos.nucleo = (SELECT nucleo FROM colaboradores WHERE id = registros_ponto.colaborador_id)
                WHERE colaborador_id = ?
                GROUP BY registros_ponto.id
                ORDER BY registros_ponto.id DESC
            """, (
                colaborador_id,
            ))
        registros = cursor.fetchall()

        cursor.execute("""
            SELECT nome
            FROM colaboradores
            WHERE id = ?
        """, (
            colaborador_id,
        ))

        nome_colaborador = cursor.fetchone()[0]
        todos = False

    conn.close()

    template_name = 'administrador/relatorios.html'
    if 'gestor_id' in session:
        template_name = 'gestor_nucleo/relatorios.html'

    return render_template(
        template_name,
        colaboradores=colaboradores,
        registros=registros,
        nome_colaborador=nome_colaborador,
        mes=mes,
        todos=todos,
        colaborador_id=colaborador_id,
        status_folhas=status_folhas
    )

# =========================
# SALVAR COLABORADOR
# =========================

@app.route('/salvar-colaborador', methods=['POST'])
def salvar_colaborador():

    conn = conectar()
    cursor = conn.cursor()

    # Se for Gestor do Núcleo, usa automaticamente o núcleo dele
    if 'gestor_id' in session:
        nucleo = session['nucleo_gestor']

    # Se for Administrador, permite escolher o núcleo
    else:
        nucleo = request.form['nucleo']

    cursor.execute("""
        INSERT INTO colaboradores (
            nome,
            vinculo,
            nucleo,
            turno,
            horario,
            presencial,
            unidade_exercicio,
            procurador_monitor,
            celular,
            login,
            senha
        )
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
    """, (
        request.form['nome'],
        request.form['vinculo'],
        nucleo,
        request.form['turno'],
        request.form['horario'],
        request.form['presencial'],
        request.form['unidade_exercicio'],
        request.form['procurador_monitor'],
        request.form['celular'],
        request.form['login'],
        request.form['senha']
    ))

    conn.commit()
    conn.close()

    return redirect('/colaboradores-cadastrados')


# =========================
# CANCELAR COLABORADOR
# =========================

@app.route('/cancelar-colaborador/<int:id>', methods=['POST'])
def cancelar_colaborador(id):

    motivo = request.form.get('motivo_cancelamento', '').strip()

    conn = conectar()
    cursor = conn.cursor()

    cursor.execute("""
        UPDATE colaboradores
        SET status = 'Inativo', cancel_observacao = ?
        WHERE id = ?
    """, (
        motivo,
        id,
    ))

    conn.commit()
    conn.close()

    return redirect('/colaboradores-cadastrados')


@app.route('/recadastrar-colaborador/<int:id>', methods=['POST'])
def recadastrar_colaborador(id):

    conn = conectar()
    cursor = conn.cursor()

    cursor.execute("""
        UPDATE colaboradores
        SET status = 'Ativo', cancel_observacao = NULL
        WHERE id = ?
    """, (
        id,
    ))

    conn.commit()
    conn.close()

    return redirect('/colaboradores-cadastrados')


# =========================
# EDITAR COLABORADOR
# =========================

@app.route('/editar-colaborador/<int:id>')
def editar_colaborador(id):

    conn = conectar()
    cursor = conn.cursor()

    cursor.execute("""
        SELECT *
        FROM colaboradores
        WHERE id = ?
    """, (
        id,
    ))

    colaborador = cursor.fetchone()

    conn.close()

    return render_template(
        'administrador/editar_colaborador.html',
        colaborador=colaborador
    )


# =========================
# ATUALIZAR COLABORADOR
# =========================

@app.route('/atualizar-colaborador/<int:id>', methods=['POST'])
def atualizar_colaborador(id):

    conn = conectar()
    cursor = conn.cursor()

    cursor.execute("""
        UPDATE colaboradores
        SET
            nome = ?,
            nucleo = ?,
            turno = ?,
            horario = ?,
            celular = ?
        WHERE id = ?
    """, (
        request.form['nome'],
        request.form['nucleo'],
        request.form['turno'],
        request.form['horario'],
        request.form['celular'],
        id
    ))

    conn.commit()
    conn.close()

    return redirect('/colaboradores-cadastrados')


# =========================
# REGISTRAR ENTRADA
# =========================

@app.route('/registrar-entrada', methods=['POST'])
def registrar_entrada():

    if 'colaborador_id' not in session:
        return redirect('/login-colaborador')

    conn = conectar()
    cursor = conn.cursor()

    hora_atual = request.form['entrada']

    data_hoje = datetime.now(
        ZoneInfo("America/Sao_Paulo")
    ).strftime('%d/%m/%Y')

    cursor.execute("""
        SELECT id, entrada, saida_final
        FROM registros_ponto
        WHERE colaborador_id = ?
        AND data = ?
    """, (session['colaborador_id'], data_hoje))

    registro = cursor.fetchone()

    if not registro:
        cursor.execute("""
            INSERT INTO registros_ponto (
                colaborador_id,
                data,
                entrada
            )
            VALUES (?, ?, ?)
        """, (
            session['colaborador_id'],
            data_hoje,
            hora_atual
        ))
        conn.commit()
    else:
        registro_id, entrada_salva, saida_salva = registro
        if not (entrada_salva or '').strip():
            cursor.execute("""
                UPDATE registros_ponto
                SET entrada = ?
                WHERE id = ?
            """, (
                hora_atual,
                registro_id
            ))
            conn.commit()

    conn.close()

    return redirect('/ponto')


# =========================
# REGISTRAR SAÍDA FINAL
# =========================

@app.route('/registrar-saida-final', methods=['POST'])
def registrar_saida_final():

    if 'colaborador_id' not in session:
        return redirect('/login-colaborador')

    conn = conectar()
    cursor = conn.cursor()

    hora_atual = request.form['saida_final']

    data_hoje = datetime.now(
        ZoneInfo("America/Sao_Paulo")
    ).strftime('%d/%m/%Y')

    cursor.execute("""
        SELECT id, entrada, saida_final
        FROM registros_ponto
        WHERE colaborador_id = ?
        AND data = ?
    """, (
        session['colaborador_id'],
        data_hoje
    ))

    registro = cursor.fetchone()

    if registro:
        registro_id, entrada_salva, saida_salva = registro
        if not (saida_salva or '').strip():
            cursor.execute("""
                UPDATE registros_ponto
                SET saida_final = ?
                WHERE id = ?
            """, (
                hora_atual,
                registro_id
            ))
            conn.commit()

    conn.close()

    return redirect('/ponto')

@app.route('/meu-relatorio')
def meu_relatorio():

    if 'colaborador_id' not in session:
        return redirect('/login-colaborador')

    from database.conexao import conectar

    conn = conectar()
    cursor = conn.cursor()

    mes_selecionado = request.args.get('mes')

    if mes_selecionado:

        ano = mes_selecionado.split('-')[0]
        mes = mes_selecionado.split('-')[1]

        cursor.execute("""
            SELECT
                registros_ponto.data,
                registros_ponto.entrada,
                registros_ponto.saida_final,
                COALESCE(NULLIF(registros_ponto.observacao, ''), group_concat(eventos.tipo || ' - ' || eventos.titulo, '; ')) AS observacao
            FROM registros_ponto
            LEFT JOIN colaboradores ON colaboradores.id = registros_ponto.colaborador_id
            LEFT JOIN eventos ON eventos.data = registros_ponto.data
                AND eventos.nucleo = colaboradores.nucleo
            WHERE colaborador_id = ?
            AND substr(registros_ponto.data,4,2) = ?
            AND substr(registros_ponto.data,7,4) = ?
            GROUP BY registros_ponto.id
            ORDER BY registros_ponto.id DESC
        """, (
            session['colaborador_id'],
            mes,
            ano
        ))

    else:

        cursor.execute("""
            SELECT
                registros_ponto.data,
                registros_ponto.entrada,
                registros_ponto.saida_final,
                COALESCE(NULLIF(registros_ponto.observacao, ''), group_concat(eventos.tipo || ' - ' || eventos.titulo, '; ')) AS observacao
            FROM registros_ponto
            LEFT JOIN colaboradores ON colaboradores.id = registros_ponto.colaborador_id
            LEFT JOIN eventos ON eventos.data = registros_ponto.data
                AND eventos.nucleo = colaboradores.nucleo
            WHERE colaborador_id = ?
            GROUP BY registros_ponto.id
            ORDER BY registros_ponto.id DESC
        """, (session['colaborador_id'],))

    registros = cursor.fetchall()

    cursor.execute("""
        SELECT folha_assinada_nome
        FROM colaboradores
        WHERE id = ?
    """, (session['colaborador_id'],))
    folha_assinada = cursor.fetchone()

    conn.close()

    return render_template(
        'colaborador/meu_relatorio.html',
        registros=registros,
        nome=session['nome'],
        mes_selecionado=mes_selecionado,
        folha_assinada_nome=folha_assinada[0] if folha_assinada else None
    )


@app.route('/anexar-folha-assinada', methods=['POST'])
def anexar_folha_assinada():

    if 'colaborador_id' not in session:
        return redirect('/login-colaborador')

    arquivo = request.files.get('arquivo_folha')
    if not arquivo or arquivo.filename == '':
        return redirect('/meu-relatorio')

    if not arquivo.filename.lower().endswith('.pdf'):
        return redirect('/meu-relatorio')

    pasta_upload = os.path.abspath('uploads/folhas_assinadas')
    os.makedirs(pasta_upload, exist_ok=True)

    nome_seguro = secure_filename(arquivo.filename)
    nome_arquivo = f"{session['colaborador_id']}_{nome_seguro}"
    caminho_arquivo = os.path.join(pasta_upload, nome_arquivo)
    arquivo.save(caminho_arquivo)

    conn = conectar()
    cursor = conn.cursor()
    cursor.execute("""
        UPDATE colaboradores
        SET folha_assinada_path = ?, folha_assinada_nome = ?
        WHERE id = ?
    """, (
        caminho_arquivo,
        nome_seguro,
        session['colaborador_id']
    ))
    conn.commit()
    conn.close()

    return redirect('/meu-relatorio')


@app.route('/download-folhas-assinadas')
def download_folhas_assinadas():

    if 'administrador_id' not in session and 'gestor_id' not in session:
        return redirect('/login-gestor-nucleo')

    conn = conectar()
    cursor = conn.cursor()

    if 'gestor_id' in session:
        cursor.execute("""
            SELECT id, nome, folha_assinada_path
            FROM colaboradores
            WHERE status = 'Ativo'
            AND nucleo = ?
            AND folha_assinada_path IS NOT NULL
            AND trim(folha_assinada_path) != ''
            ORDER BY nome
        """, (session['nucleo_gestor'],))
    else:
        cursor.execute("""
            SELECT id, nome, folha_assinada_path
            FROM colaboradores
            WHERE status = 'Ativo'
            AND folha_assinada_path IS NOT NULL
            AND trim(folha_assinada_path) != ''
            ORDER BY nome
        """)

    anexos = cursor.fetchall()
    conn.close()

    if not anexos:
        return redirect('/relatorios')

    pasta_upload = os.path.abspath('uploads/folhas_assinadas')
    os.makedirs(pasta_upload, exist_ok=True)

    nome_zip = os.path.join(pasta_upload, f"pacote_folhas_assinadas_{datetime.now().strftime('%Y%m%d_%H%M%S')}.zip")

    with ZipFile(nome_zip, 'w') as zip_file:
        for colaborador_id, nome, caminho in anexos:
            if os.path.exists(caminho):
                nome_arquivo_zip = f"{nome.replace('/', '_').replace('\\', '_')}_{colaborador_id}.pdf"
                zip_file.write(caminho, arcname=nome_arquivo_zip)

    return send_file(
        nome_zip,
        as_attachment=True,
        download_name='pacote_folhas_assinadas.zip'
    )


@app.route('/salvar-observacao', methods=['POST'])
def salvar_observacao():

    if 'colaborador_id' not in session:
        return redirect('/login-colaborador')

    from database.conexao import conectar
    from datetime import datetime

    conn = conectar()
    cursor = conn.cursor()

    data = request.form['data']
    observacao = request.form['observacao']

    data_convertida = datetime.strptime(
        data,
        "%Y-%m-%d"
    ).strftime("%d/%m/%Y")

    cursor.execute("""
        UPDATE registros_ponto
        SET observacao = ?
        WHERE colaborador_id = ?
        AND data = ?
    """, (
        observacao,
        session['colaborador_id'],
        data_convertida
    ))

    conn.commit()
    conn.close()

    return redirect('/meu-relatorio')

from flask import send_file
from reportlab.pdfgen import canvas

@app.route('/gerar-pdf')
def gerar_pdf():

    colaborador_id = request.args.get('colaborador_id')
    mes_selecionado = request.args.get('mes')

    if 'administrador_id' not in session and 'gestor_id' not in session and 'colaborador_id' not in session:
        return redirect('/login-colaborador')

    from database.conexao import conectar

    conn = conectar()
    cursor = conn.cursor()

    if 'administrador_id' in session or 'gestor_id' in session:
        if not colaborador_id:
            conn.close()
            return redirect('/relatorios')

        if colaborador_id == 'all':
            if 'administrador_id' not in session:
                conn.close()
                return redirect('/relatorios')

            colaborador = ('Todos os Colaboradores', '', '')
        else:
            cursor.execute("""
                SELECT
                    nome,
                    nucleo,
                    horario
                FROM colaboradores
                WHERE id = ?
            """, (colaborador_id,))
            colaborador = cursor.fetchone()

            if not colaborador:
                conn.close()
                return redirect('/relatorios')

            if 'gestor_id' in session:
                cursor.execute("""
                    SELECT id
                    FROM colaboradores
                    WHERE id = ?
                    AND nucleo = ?
                """, (colaborador_id, session['nucleo_gestor']))
                if not cursor.fetchone():
                    conn.close()
                    return redirect('/relatorios')
    else:
        colaborador_id = session['colaborador_id']
        cursor.execute("""
            SELECT
                nome,
                nucleo,
                horario
            FROM colaboradores
            WHERE id = ?
        """, (colaborador_id,))
        colaborador = cursor.fetchone()

    print("MES PDF:", mes_selecionado)

    if colaborador_id == 'all':
        if mes_selecionado and "-" in mes_selecionado:
            ano, mes = mes_selecionado.split('-')
            cursor.execute("""
                SELECT
                    registros_ponto.data,
                    registros_ponto.entrada,
                    registros_ponto.saida_final,
                    registros_ponto.observacao,
                    colaboradores.nome
                FROM registros_ponto
                INNER JOIN colaboradores ON colaboradores.id = registros_ponto.colaborador_id
                WHERE substr(registros_ponto.data,4,2) = ?
                AND substr(registros_ponto.data,7,4) = ?
                ORDER BY registros_ponto.data
            """, (
                mes,
                ano
            ))
        else:
            cursor.execute("""
                SELECT
                    registros_ponto.data,
                    registros_ponto.entrada,
                    registros_ponto.saida_final,
                    registros_ponto.observacao,
                    colaboradores.nome
                FROM registros_ponto
                INNER JOIN colaboradores ON colaboradores.id = registros_ponto.colaborador_id
                ORDER BY registros_ponto.data
            """)
        registros = cursor.fetchall()
        nome_colaborador = colaborador[0].replace(" ", "_")

        if mes_selecionado and "-" in mes_selecionado:
            ano = mes_selecionado.split('-')[0]
            mes = mes_selecionado.split('-')[1]
            arquivo = f"Folha_Frequencia_Todos_Colaboradores_{mes}_{ano}.pdf"
        else:
            arquivo = f"Folha_Frequencia_Todos_Colaboradores.pdf"
    else:
        if mes_selecionado and "-" in mes_selecionado:
            ano, mes = mes_selecionado.split('-')
            cursor.execute("""
                SELECT
                    data,
                    entrada,
                    saida_final,
                    observacao
                FROM registros_ponto
                WHERE colaborador_id = ?
                AND substr(data,4,2) = ?
                AND substr(data,7,4) = ?
                ORDER BY data
            """, (
                colaborador_id,
                mes,
                ano
            ))
        else:
            cursor.execute("""
                SELECT
                    data,
                    entrada,
                    saida_final,
                    observacao
                FROM registros_ponto
                WHERE colaborador_id = ?
                ORDER BY data
            """, (colaborador_id,))
        registros = cursor.fetchall()
        nome_colaborador = colaborador[0].replace(" ", "_")

        if mes_selecionado and "-" in mes_selecionado:
            ano = mes_selecionado.split('-')[0]
            mes = mes_selecionado.split('-')[1]
            arquivo = f"Folha_Frequencia_{nome_colaborador}_{mes}_{ano}.pdf"
        else:
            arquivo = f"Folha_Frequencia_{nome_colaborador}.pdf"

    if colaborador_id == 'all':
        doc = SimpleDocTemplate(arquivo)
        estilos = getSampleStyleSheet()
        elementos = []
        elementos.append(
            Paragraph(
                "<b>FOLHA DE FREQUÊNCIA</b>",
                estilos['Title']
            )
        )
        if mes_selecionado and "-" in mes_selecionado:
            elementos.append(
                Paragraph(
                    f"<b>Competência:</b> {mes}/{ano}",
                    estilos['Normal']
                )
            )
            elementos.append(Spacer(1, 12))
        elementos.append(
            Paragraph(
                f"<b>Relatório:</b> Todos os colaboradores",
                estilos['Normal']
            )
        )
        elementos.append(Spacer(1, 20))
        dados = [
            ["Data", "Entrada", "Saída", "Observação", "Colaborador"]
        ]
        for registro in registros:
            dados.append([
                registro[0],
                registro[1] or "",
                registro[2] or "",
                registro[3] or "",
                registro[4] or ""
            ])
        tabela = Table(dados)
        tabela.setStyle(TableStyle([
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
        ]))
        elementos.append(tabela)
        elementos.append(Spacer(1, 120))
        elementos.append(
            Paragraph(
                "<b>VISTO</b>",
                estilos['Title']
            )
        )
        elementos.append(Spacer(1, 50))
        assinaturas = Table([
            [
                "__________________________________",
                "__________________________________"
            ],
            [
                colaborador[0],
                "ASSINATURA/CARIMBO"
            ],
            [
                "Relatório Abrangente",
                "PROCURADOR DO ESTADO"
            ]
        ], colWidths=[250, 250])
        assinaturas.setStyle(TableStyle([
            ('ALIGN', (0,0), (-1,-1), 'CENTER'),
            ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
            ('FONTNAME', (0,0), (-1,-1), 'Helvetica'),
            ('BOTTOMPADDING', (0,0), (-1,-1), 8),
        ]))
        elementos.append(assinaturas)
        doc.build(elementos)
    else:
        doc = SimpleDocTemplate(arquivo)
        estilos = getSampleStyleSheet()
        elementos = []
        elementos.append(
            Paragraph(
                "<b>FOLHA DE FREQUÊNCIA</b>",
                estilos['Title']
            )
        )
        if mes_selecionado and "-" in mes_selecionado:
            elementos.append(
                Paragraph(
                    f"<b>Competência:</b> {mes}/{ano}",
                    estilos['Normal']
                )
            )
            elementos.append(Spacer(1, 12))
        elementos.append(
            Paragraph(
                f"<b>Nome:</b> {colaborador[0]}",
                estilos['Normal']
            )
        )
        elementos.append(
            Paragraph(
                f"<b>Núcleo:</b> {colaborador[1]}",
                estilos['Normal']
            )
        )
        elementos.append(
            Paragraph(
                f"<b>Horário:</b> {colaborador[2]}",
                estilos['Normal']
            )
        )
        elementos.append(Spacer(1, 20))
        dados = [
            ["Data", "Entrada", "Saída", "Observação"]
        ]
        for registro in registros:
            dados.append([
                registro[0],
                registro[1] or "",
                registro[2] or "",
                registro[3] or ""
            ])
        tabela = Table(dados)
        tabela.setStyle(TableStyle([
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
        ]))
        elementos.append(tabela)
        elementos.append(Spacer(1, 120))
        elementos.append(
            Paragraph(
                "<b>VISTO</b>",
                estilos['Title']
            )
        )
        elementos.append(Spacer(1, 50))
        assinaturas = Table([
            [
                "__________________________________",
                "__________________________________"
            ],
            [
                colaborador[0],
                "ASSINATURA/CARIMBO"
            ],
            [
                "Assinatura do Colaborador",
                "PROCURADOR DO ESTADO"
            ]
        ], colWidths=[250, 250])
        assinaturas.setStyle(TableStyle([
            ('ALIGN', (0,0), (-1,-1), 'CENTER'),
            ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
            ('FONTNAME', (0,0), (-1,-1), 'Helvetica'),
            ('BOTTOMPADDING', (0,0), (-1,-1), 8),
        ]))
        elementos.append(assinaturas)
        doc.build(elementos)

    return send_file(
        arquivo,
        as_attachment=True
    )

@app.route('/gerar-excel')
def gerar_excel():

    colaborador_id = request.args.get('colaborador_id')
    mes_selecionado = request.args.get('mes')

    if 'administrador_id' not in session and 'gestor_id' not in session and 'colaborador_id' not in session:
        return redirect('/login-colaborador')

    from database.conexao import conectar

    conn = conectar()
    cursor = conn.cursor()

    if 'administrador_id' in session or 'gestor_id' in session:
        if not colaborador_id:
            conn.close()
            return redirect('/relatorios')

        if colaborador_id == 'all':
            if 'administrador_id' not in session:
                conn.close()
                return redirect('/relatorios')
            colaborador = ('Todos os Colaboradores', '', '')
        else:
            cursor.execute("""
                SELECT
                    nome,
                    nucleo,
                    horario
                FROM colaboradores
                WHERE id = ?
            """, (colaborador_id,))
            colaborador = cursor.fetchone()
            if not colaborador:
                conn.close()
                return redirect('/relatorios')
            if 'gestor_id' in session:
                cursor.execute("""
                    SELECT id
                    FROM colaboradores
                    WHERE id = ?
                    AND nucleo = ?
                """, (colaborador_id, session['nucleo_gestor']))
                if not cursor.fetchone():
                    conn.close()
                    return redirect('/relatorios')
    else:
        colaborador_id = session['colaborador_id']
        cursor.execute("""
            SELECT
                nome,
                nucleo,
                horario
            FROM colaboradores
            WHERE id = ?
        """, (colaborador_id,))
        colaborador = cursor.fetchone()

    print("MES EXCEL:", mes_selecionado)

    if colaborador_id == 'all':
        if mes_selecionado and "-" in mes_selecionado:
            ano, mes = mes_selecionado.split('-')
            cursor.execute("""
                SELECT
                    data,
                    entrada,
                    saida_final,
                    observacao,
                    colaboradores.nome
                FROM registros_ponto
                INNER JOIN colaboradores ON colaboradores.id = registros_ponto.colaborador_id
                WHERE substr(registros_ponto.data,4,2) = ?
                AND substr(registros_ponto.data,7,4) = ?
                ORDER BY data
            """, (
                mes,
                ano
            ))
        else:
            cursor.execute("""
                SELECT
                    data,
                    entrada,
                    saida_final,
                    observacao,
                    colaboradores.nome
                FROM registros_ponto
                INNER JOIN colaboradores ON colaboradores.id = registros_ponto.colaborador_id
                ORDER BY data
            """)
    else:
        if mes_selecionado and "-" in mes_selecionado:
            ano, mes = mes_selecionado.split('-')
            cursor.execute("""
                SELECT
                    data,
                    entrada,
                    saida_final,
                    observacao
                FROM registros_ponto
                WHERE colaborador_id = ?
                AND substr(data,4,2) = ?
                AND substr(data,7,4) = ?
                ORDER BY data
            """, (
                colaborador_id,
                mes,
                ano
            ))
        else:
            cursor.execute("""
                SELECT
                    data,
                    entrada,
                    saida_final,
                    observacao
                FROM registros_ponto
                WHERE colaborador_id = ?
                ORDER BY data
            """, (colaborador_id,))

    registros = cursor.fetchall()
    conn.close()

    wb = Workbook()
    ws = wb.active
    ws.title = "Frequência"
    ws['A1'] = "FOLHA DE FREQUÊNCIA"

    if colaborador_id == 'all':
        if mes_selecionado and "-" in mes_selecionado:
            ano = mes_selecionado.split('-')[0]
            mes = mes_selecionado.split('-')[1]
            ws['A2'] = f"Competência: {mes}/{ano}"
        ws['A3'] = "Relatório"
        ws['B3'] = "Todos os Colaboradores"
        ws['A4'] = "Núcleo"
        ws['B4'] = "Todos"
        ws['A5'] = "Horário"
        ws['B5'] = "Todos"
        ws.append([])
        ws.append([
            "Data",
            "Entrada",
            "Saída",
            "Observação",
            "Colaborador"
        ])
        for registro in registros:
            ws.append([
                registro[0],
                registro[1] or "",
                registro[2] or "",
                registro[3] or "",
                registro[4] or ""
            ])
        if mes_selecionado and "-" in mes_selecionado:
            arquivo = f"Folha_Frequencia_Todos_Colaboradores_{mes}_{ano}.xlsx"
        else:
            arquivo = f"Folha_Frequencia_Todos_Colaboradores.xlsx"
    else:
        if mes_selecionado and "-" in mes_selecionado:
            ano = mes_selecionado.split('-')[0]
            mes = mes_selecionado.split('-')[1]
            ws['A2'] = f"Competência: {mes}/{ano}"
        ws['A3'] = "Nome"
        ws['B3'] = colaborador[0]
        ws['A4'] = "Núcleo"
        ws['B4'] = colaborador[1]
        ws['A5'] = "Horário"
        ws['B5'] = colaborador[2]
        ws.append([])
        ws.append([
            "Data",
            "Entrada",
            "Saída",
            "Observação"
        ])
        for registro in registros:
            ws.append([
                registro[0],
                registro[1] or "",
                registro[2] or "",
                registro[3] or ""
            ])
        nome_colaborador = colaborador[0].replace(" ", "_")
        if mes_selecionado and "-" in mes_selecionado:
            arquivo = f"Folha_Frequencia_{nome_colaborador}_{mes}_{ano}.xlsx"
        else:
            arquivo = f"Folha_Frequencia_{nome_colaborador}.xlsx"

    wb.save(arquivo)

    return send_file(
        arquivo,
        as_attachment=True
    )

@app.route('/primeiro-acesso')
def primeiro_acesso():

    if 'colaborador_id' not in session:
        return redirect('/login-colaborador')

    return render_template(
    'colaborador/primeiro_acesso.html'
)

@app.route('/alterar-senha', methods=['POST'])
def alterar_senha():

    if 'colaborador_id' not in session:
        return redirect('/login-colaborador')

    senha1 = request.form['senha']
    senha2 = request.form['confirmar_senha']

    if senha1 != senha2:
        return """
        As senhas não coincidem.
        """

    if senha1 == "Novocolab123":
        return """
        A senha padrão não pode ser utilizada novamente.
        """

    padrao = r'^(?=.*[a-z])(?=.*[A-Z])(?=.*\d)(?=.*[@$!%*?&]).{8,}$'

    if not re.match(padrao, senha1):

        return """
        A senha deve possuir:

        - mínimo de 8 caracteres;
        - uma letra maiúscula;
        - uma letra minúscula;
        - um número;
        - um caractere especial.
        """

    from database.conexao import conectar

    conn = conectar()
    cursor = conn.cursor()

    cursor.execute("""
        UPDATE colaboradores
        SET senha = ?
        WHERE id = ?
    """, (
        senha1,
        session['colaborador_id']
    ))

    conn.commit()
    conn.close()

    return redirect('/ponto')

@app.route('/lancar-ponto')
def lancar_ponto():

    if 'administrador_id' not in session:
        return redirect('/login-administrador')

    conn = conectar()
    cursor = conn.cursor()

    cursor.execute("""
        SELECT id, nome
        FROM colaboradores
        WHERE status = 'Ativo'
        ORDER BY nome
    """)

    colaboradores = cursor.fetchall()

    conn.close()

    return render_template(
        'administrador/lancar_ponto.html',
        colaboradores=colaboradores
    )

@app.route('/salvar-lancamento-ponto', methods=['POST'])
def salvar_lancamento_ponto():

    if 'administrador_id' not in session:
        return redirect('/login-administrador')

    colaborador_id = request.form['colaborador_id']
    data = request.form['data']
    entrada = request.form['entrada']
    saida_final = request.form['saida_final']
    observacao = request.form['observacao']

    conn = conectar()
    cursor = conn.cursor()

    cursor.execute("""
        INSERT INTO registros_ponto (
            colaborador_id,
            data,
            entrada,
            saida_final,
            observacao
        )
        VALUES (?, ?, ?, ?, ?)
    """, (
        colaborador_id,
        data,
        entrada,
        saida_final,
        observacao
    ))

    conn.commit()
    conn.close()

    return redirect('/lancar-ponto')

@app.route('/enviar-ajuste', methods=['POST'])
def enviar_ajuste():

    if 'colaborador_id' not in session:
        return redirect('/login-colaborador')

    conn = conectar()
    cursor = conn.cursor()

    cursor.execute("""
        INSERT INTO ajustes_ponto (
            colaborador_id,
            data,
            motivo,
            status
        )
        VALUES (?, ?, ?, ?)
    """, (
        session['colaborador_id'],
        request.form['data'],
        request.form['motivo'],
        'Pendente'
    ))

    conn.commit()
    conn.close()

    return redirect('/meus-ajustes')

@app.route('/primeiro-acesso-gestor')
def primeiro_acesso_gestor():

    if 'gestor_id' not in session:
        return redirect('/login-gestor-nucleo')

    return render_template(
        'gestor_nucleo/primeiro_acesso_gestor.html'
    )

@app.route('/alterar-senha-gestor', methods=['POST'])
def alterar_senha_gestor():

    if 'gestor_id' not in session:
        return redirect('/login-gestor-nucleo')

    senha1 = request.form['senha']
    senha2 = request.form['confirmar_senha']

    if senha1 != senha2:
        return "As senhas não coincidem."

    if senha1 == "Novocolab123":
        return "A senha padrão não pode ser utilizada novamente."

    padrao = r'^(?=.*[a-z])(?=.*[A-Z])(?=.*\d)(?=.*[@$!%*?&]).{8,}$'

    if not re.match(padrao, senha1):

        return """
        A senha deve possuir:

        - mínimo de 8 caracteres;
        - uma letra maiúscula;
        - uma letra minúscula;
        - um número;
        - um caractere especial.
        """

    conn = conectar()
    cursor = conn.cursor()

    cursor.execute("""
        UPDATE gestores
        SET senha = ?
        WHERE id = ?
    """, (
        senha1,
        session['gestor_id']
    ))

    conn.commit()
    conn.close()

    return redirect('/dashboard-gestor-nucleo')

if __name__ == '__main__':
    app.run(debug=True)