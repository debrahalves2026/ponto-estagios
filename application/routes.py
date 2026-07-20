from flask import Blueprint, render_template, request, redirect, session, send_file, after_this_request
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from openpyxl import Workbook
from openpyxl.styles import Alignment
from datetime import datetime, timedelta
from zoneinfo import ZoneInfo
from werkzeug.utils import secure_filename
from zipfile import ZipFile
from urllib.parse import quote_plus
import os
import re
import tempfile
from xml.sax.saxutils import escape
from .app_utils import (
    get_db_connection,
    validate_password,
    login_required,
    ensure_upload_folder,
    save_temp_file,
    clean_temp_file,
    safe_filename,
    db_cursor
)

def _celula_pdf(valor, estilo):
    """Prepara o texto para células da tabela PDF com quebra automática."""
    if valor is None:
        valor = ""
    return Paragraph(escape(str(valor)), estilo)


main_bp = Blueprint('main', __name__)


def _usuario_log_atual():
    if 'administrador_id' in session:
        return 'Administrador', session.get('nome_admin', 'Administrador')
    if 'gestor_id' in session:
        return 'Gestor', session.get('nome_gestor', 'Gestor')
    if 'colaborador_id' in session:
        return 'Colaborador', session.get('nome', 'Colaborador')
    return 'Sistema', 'Sistema'


def _registrar_log(cursor, acao, detalhes):
    tipo_usuario, nome_usuario = _usuario_log_atual()
    data_hora = datetime.now(
        ZoneInfo("America/Sao_Paulo")
    ).strftime('%d/%m/%Y %H:%M:%S')
    cursor.execute("""
        INSERT INTO logs_sistema (
            data_hora,
            tipo_usuario,
            nome_usuario,
            acao,
            detalhes
        )
        VALUES (?, ?, ?, ?, ?)
    """, (
        data_hora,
        tipo_usuario,
        nome_usuario,
        acao,
        detalhes
    ))


def _agora_sp():
    return datetime.now(ZoneInfo("America/Sao_Paulo"))


def _data_iso_para_br(data_iso):
    try:
        return datetime.strptime(data_iso, '%Y-%m-%d').strftime('%d/%m/%Y')
    except (ValueError, TypeError):
        return data_iso


def _data_br_para_date(data_br):
    return datetime.strptime(data_br, '%d/%m/%Y').date()


def _eh_fim_de_semana(data_br):
    return _data_br_para_date(data_br).weekday() >= 5


def _nome_reprovador():
    return session.get('nome_admin') or session.get('nome_gestor', 'Usuário')


def _mensagem_query(param):
    return request.args.get(param, '').strip()


def _redirect_com_mensagem(path, chave, mensagem):
    return redirect(f"{path}?{chave}={quote_plus(mensagem)}")


def _colaborador_nucleo(cursor, colaborador_id):
    cursor.execute("""
        SELECT nome, nucleo
        FROM colaboradores
        WHERE id = ?
    """, (colaborador_id,))
    return cursor.fetchone()


def _eh_feriado(cursor, data_br, nucleo):
    cursor.execute("""
        SELECT 1
        FROM eventos
        WHERE data = ?
        AND lower(trim(tipo)) = 'feriado'
        AND (nucleo = ? OR nucleo IS NULL OR trim(nucleo) = '')
        LIMIT 1
    """, (
        data_br,
        nucleo
    ))
    return cursor.fetchone() is not None


def _datas_feriado_nucleo(cursor, nucleo):
    cursor.execute("""
        SELECT data
        FROM eventos
        WHERE lower(trim(tipo)) = 'feriado'
        AND (nucleo = ? OR nucleo IS NULL OR trim(nucleo) = '')
    """, (nucleo,))

    datas_iso = []
    for row in cursor.fetchall():
        data_br = (row[0] or '').strip()
        try:
            datas_iso.append(datetime.strptime(data_br, '%d/%m/%Y').strftime('%Y-%m-%d'))
        except ValueError:
            continue

    return sorted(set(datas_iso))


def _bloqueio_trabalho(cursor, data_br, nucleo):
    if _eh_fim_de_semana(data_br):
        return 'Não é permitido registrar ou solicitar ajuste para finais de semana.'
    if _eh_feriado(cursor, data_br, nucleo):
        return 'Não é permitido registrar ou solicitar ajuste para feriados.'
    return None


def _proximo_dia_util(cursor, nucleo, data_inicial=None):
    data_atual = data_inicial or _agora_sp().date()

    for _ in range(370):
        data_br = data_atual.strftime('%d/%m/%Y')
        if not _bloqueio_trabalho(cursor, data_br, nucleo):
            return data_atual
        data_atual += timedelta(days=1)

    return data_inicial or _agora_sp().date()


def _registros_folha_colaborador(cursor, colaborador_id, mes=None, ano=None, ordem='desc'):
    cursor.execute("""
        SELECT nucleo
        FROM colaboradores
        WHERE id = ?
    """, (colaborador_id,))
    colaborador = cursor.fetchone()
    nucleo = (colaborador[0] if colaborador else '') or ''

    filtro_periodo = ''
    parametros_registros = [colaborador_id]
    parametros_eventos = [nucleo]

    if mes and ano:
        filtro_periodo = " AND substr(data,4,2) = ? AND substr(data,7,4) = ?"
        parametros_registros.extend([mes, ano])
        parametros_eventos.extend([mes, ano])

    cursor.execute(f"""
        SELECT
            data,
            entrada,
            saida_final,
            observacao
        FROM registros_ponto
        WHERE colaborador_id = ?
        {filtro_periodo}
    """, parametros_registros)
    registros = cursor.fetchall()

    cursor.execute(f"""
        SELECT
            data,
            group_concat(tipo || ' - ' || titulo, '; ') AS eventos_obs
        FROM eventos
        WHERE (nucleo = ? OR nucleo IS NULL OR trim(nucleo) = '')
        {filtro_periodo}
        GROUP BY data
    """, parametros_eventos)
    eventos = cursor.fetchall()

    itens_por_data = {}
    for data, entrada, saida_final, observacao in registros:
        itens_por_data[data] = {
            'data': data,
            'entrada': entrada,
            'saida_final': saida_final,
            'observacao': observacao
        }

    for data, eventos_obs in eventos:
        item = itens_por_data.get(data)
        if item:
            if not (item['observacao'] or '').strip():
                item['observacao'] = eventos_obs
        else:
            itens_por_data[data] = {
                'data': data,
                'entrada': None,
                'saida_final': None,
                'observacao': eventos_obs
            }

    def _chave_ordenacao(item):
        try:
            return _data_br_para_date(item['data'])
        except (ValueError, TypeError):
            return datetime.min.date()

    registros_ordenados = sorted(
        itens_por_data.values(),
        key=_chave_ordenacao,
        reverse=(ordem != 'asc')
    )

    return [
        (
            item['data'],
            item['entrada'],
            item['saida_final'],
            item['observacao']
        )
        for item in registros_ordenados
    ]

@main_bp.route('/')
def inicio():
    session.clear()
    return render_template('index.html')
# ADMINISTRADOR

@main_bp.route('/login-administrador', methods=['GET', 'POST'])
def login_administrador():

    if request.method == 'POST':

        login = request.form.get('login')
        senha = request.form.get('senha')

        if login == 'gestor1' and senha == 'gestor2026':

            session.clear()
            session['administrador_id'] = 1
            session['nome_admin'] = 'Administrador'

            return redirect('/dashboard-administrador')

        return render_template(
            'administrador/login_administrador.html',
            erro='Login ou senha inválidos'
        )

    return render_template(
        'administrador/login_administrador.html'
    )
# GESTOR DO NÚCLEO

@main_bp.route('/login-gestor-nucleo', methods=['GET', 'POST'])
def login_gestor_nucleo():

    if request.method == 'POST':

        login = request.form.get('login')
        senha = request.form.get('senha')

        with db_cursor() as cursor:
            cursor.execute("""
            SELECT *
            FROM gestores
            WHERE login = ?
            AND senha = ?
            AND status = 'Ativo'
        """, (
            login,
            senha
        ))

        gestor = cursor.fetchone()
        if gestor:

            session.clear()
            session['gestor_id'] = gestor[0]
            session['nome_gestor'] = gestor[1]
            session['nucleo_gestor'] = gestor[2]

            # Primeiro acesso
            if gestor[6] == "Novocolab123":
                return redirect('/primeiro-acesso-gestor')

            return redirect('/dashboard-gestor-nucleo')

        return render_template(
            'gestor_nucleo/login_gestor_nucleo.html',
            erro='Login ou senha inválidos'
        )

    return render_template(
        'gestor_nucleo/login_gestor_nucleo.html'
    )


@main_bp.route('/dashboard-gestor-nucleo')
def dashboard_gestor_nucleo():

    if 'gestor_id' not in session:
        return redirect('/login-gestor-nucleo')

    with db_cursor() as cursor:
        # Total de colaboradores
        cursor.execute("""
            SELECT COUNT(*)
            FROM colaboradores
            WHERE nucleo = ?
        """, (
            session['nucleo_gestor'],
        ))
        total_colaboradores = cursor.fetchone()[0]

        # Colaboradores ativos
        cursor.execute("""
            SELECT COUNT(*)
            FROM colaboradores
            WHERE nucleo = ?
            AND status = 'Ativo'
        """, (
            session['nucleo_gestor'],
        ))
        colaboradores_ativos = cursor.fetchone()[0]

        # Colaboradores inativos
        cursor.execute("""
            SELECT COUNT(*)
            FROM colaboradores
            WHERE nucleo = ?
            AND status = 'Inativo'
        """, (
            session['nucleo_gestor'],
        ))
        colaboradores_inativos = cursor.fetchone()[0]

        # Ajustes pendentes
        cursor.execute("""
            SELECT COUNT(*)
            FROM ajustes_ponto
            INNER JOIN colaboradores
            ON ajustes_ponto.colaborador_id = colaboradores.id
            WHERE colaboradores.nucleo = ?
            AND ajustes_ponto.status = 'Pendente'
        """, (
            session['nucleo_gestor'],
        ))
        ajustes_pendentes = cursor.fetchone()[0]

    return render_template(
        'gestor_nucleo/dashboard_gestor_nucleo.html',
        nome=session['nome_gestor'],
        nucleo=session['nucleo_gestor'],
        total_colaboradores=total_colaboradores,
        colaboradores_ativos=colaboradores_ativos,
        colaboradores_inativos=colaboradores_inativos,
        ajustes_pendentes=ajustes_pendentes
    )
# CADASTRO DE GESTOR

@main_bp.route('/cadastro-gestor')
def cadastro_gestor():

    if 'administrador_id' not in session:
        return redirect('/login-administrador')

    return render_template(
        'administrador/cadastro_gestor.html'
    )


@main_bp.route('/salvar-gestor', methods=['POST'])
def salvar_gestor():

    if 'administrador_id' not in session:
        return redirect('/login-administrador')

    nome = request.form.get('nome', '')
    nucleo = request.form.get('nucleo', '')
    unidade_exercicio = request.form.get('unidade_exercicio', '')
    celular = request.form.get('celular', '')
    login = request.form.get('login', '')

    try:
        with db_cursor(commit=True) as cursor:
            cursor.execute("""
                INSERT INTO gestores (
                    nome,
                    nucleo,
                    unidade_exercicio,
                    celular,
                    login,
                    senha
                )
                VALUES (?, ?, ?, ?, ?, ?)
            """, (
                nome,
                nucleo,
                unidade_exercicio,
                celular,
                login,
                "Novocolab123"
            ))
            _registrar_log(
                cursor,
                'Cadastro de gestor',
                f"{session.get('nome_admin', 'Administrador')} cadastrou o gestor {nome}."
            )

        return redirect('/gestores-cadastrados')
    except Exception as e:
        error_message = 'Erro ao cadastrar gestor. Verifique se o login já existe.'
        if 'UNIQUE constraint failed' in str(e):
            error_message = 'Já existe um gestor cadastrado com esse login.'

        return render_template(
            'administrador/cadastro_gestor.html',
            erro=error_message,
            nome=nome,
            nucleo=nucleo,
            unidade_exercicio=unidade_exercicio,
            celular=celular,
            login=login
        )
# GESTORES CADASTRADOS

@main_bp.route('/editar-gestor/<int:id>')
def editar_gestor(id):

    if 'administrador_id' not in session:
        return redirect('/login-administrador')

    with db_cursor() as cursor:
        cursor.execute("""
        SELECT *
        FROM gestores
        WHERE id = ?
    """, (
        id,
    ))

    gestor = cursor.fetchone()
    return render_template(
        'administrador/editar_gestor.html',
        gestor=gestor
    )


@main_bp.route('/atualizar-gestor/<int:id>', methods=['POST'])
def atualizar_gestor(id):

    if 'administrador_id' not in session:
        return redirect('/login-administrador')

    with db_cursor(commit=True) as cursor:
        cursor.execute("""
        UPDATE gestores
        SET
            nome = ?,
            nucleo = ?,
            login = ?
        WHERE id = ?
    """, (
        request.form['nome'],
        request.form['nucleo'],
        request.form['login'],
        id
    ))

    return redirect('/gestores-cadastrados')


@main_bp.route('/cancelar-gestor/<int:id>', methods=['POST'])
def cancelar_gestor(id):

    if 'administrador_id' not in session:
        return redirect('/login-administrador')

    motivo = request.form.get('motivo_cancelamento', '').strip()

    with db_cursor(commit=True) as cursor:
        cursor.execute("""
        UPDATE gestores
        SET status = 'Inativo', cancel_observacao = ?
        WHERE id = ?
    """, (
        motivo,
        id,
    ))

    return redirect('/gestores-cadastrados')

@main_bp.route('/recadastrar-gestor/<int:id>', methods=['POST'])
def recadastrar_gestor(id):

    if 'administrador_id' not in session:
        return redirect('/login-administrador')

    with db_cursor(commit=True) as cursor:
        cursor.execute("""
        UPDATE gestores
        SET status = 'Ativo', cancel_observacao = NULL
        WHERE id = ?
    """, (
        id,
    ))

    return redirect('/gestores-cadastrados')

@main_bp.route('/resetar-senha-gestor/<int:id>')
def resetar_senha_gestor(id):

    if 'administrador_id' not in session:
        return redirect('/login-administrador')

    with db_cursor(commit=True) as cursor:
        cursor.execute("""
        UPDATE gestores
        SET senha = ?
        WHERE id = ?
    """, (
        "Novocolab123",
        id
    ))

    return redirect('/gestores-cadastrados')

@main_bp.route('/resetar-senha-colaborador/<int:id>')
def resetar_senha_colaborador(id):

    if 'administrador_id' not in session:
        return redirect('/login-administrador')

    with db_cursor(commit=True) as cursor:
        cursor.execute("""
        UPDATE colaboradores
        SET senha = ?
        WHERE id = ?
    """, (
        "Novocolab123",
        id
    ))

    return redirect('/colaboradores-cadastrados')

@main_bp.route('/gestores-cadastrados')
def gestores_cadastrados():

    if 'administrador_id' not in session:
        return redirect('/login-administrador')

    with db_cursor() as cursor:
        cursor.execute("""
        SELECT *
        FROM gestores
        ORDER BY nome
    """)

    gestores = cursor.fetchall()
    return render_template(
        'administrador/gestores_cadastrados.html',
        gestores=gestores
    )
# COLABORADOR

@main_bp.route('/login-colaborador', methods=['GET', 'POST'])
def login_colaborador():

    if request.method == 'POST':

        login = request.form.get('login')
        senha = request.form.get('senha')

        with db_cursor() as cursor:
            cursor.execute("""
            SELECT *
            FROM colaboradores
            WHERE login = ?
            AND senha = ?
            AND status = 'Ativo'
        """, (
            login,
            senha
        ))

        colaborador = cursor.fetchone()
        if colaborador:

            session.clear()
            session['colaborador_id'] = colaborador[0]
            session['nome'] = colaborador[1]

            if colaborador[11] == "Novocolab123":
                return redirect('/primeiro-acesso')

            return redirect('/ponto')

        return render_template(
            'colaborador/login_colaborador.html',
            erro='Login ou senha inválidos'
        )

    return render_template(
        'colaborador/login_colaborador.html'
    )
# CADASTRO DE COLABORADOR

@main_bp.route('/cadastro-colaborador')
def cadastro_colaborador():

    if 'administrador_id' in session:
        return render_template(
            'administrador/cadastro_colaborador.html'
        )

    if 'gestor_id' in session:
        return render_template(
            'gestor_nucleo/cadastro_colaborador.html',
            nucleo=session['nucleo_gestor']
        )

    return redirect('/')
# DASHBOARD ADMINISTRADOR

@main_bp.route('/dashboard-administrador')
def dashboard_administrador():

    with db_cursor() as cursor:
        # Total de colaboradores
        cursor.execute("""
            SELECT COUNT(*)
            FROM colaboradores
        """)
        total_colaboradores = cursor.fetchone()[0]

    # Colaboradores ativos
    cursor.execute("""
        SELECT COUNT(*)
        FROM colaboradores
        WHERE status = 'Ativo'
    """)
    colaboradores_ativos = cursor.fetchone()[0]

    # Colaboradores inativos
    cursor.execute("""
        SELECT COUNT(*)
        FROM colaboradores
        WHERE status = 'Inativo'
    """)
    colaboradores_inativos = cursor.fetchone()[0]

    # Total de gestores
    cursor.execute("""
        SELECT COUNT(*)
        FROM gestores
        WHERE status = 'Ativo'
    """)
    total_gestores = cursor.fetchone()[0]

    # Total de núcleos
    cursor.execute("""
        SELECT COUNT(DISTINCT nucleo)
        FROM colaboradores
    """)
    total_nucleos = cursor.fetchone()[0]

    # Ajustes pendentes
    cursor.execute("""
        SELECT COUNT(*)
        FROM ajustes_ponto
        WHERE status = 'Pendente'
    """)
    ajustes_pendentes = cursor.fetchone()[0]
    return render_template(
        'administrador/dashboard_administrador.html',
        total_colaboradores=total_colaboradores,
        colaboradores_ativos=colaboradores_ativos,
        colaboradores_inativos=colaboradores_inativos,
        total_gestores=total_gestores,
        total_nucleos=total_nucleos,
        ajustes_pendentes=ajustes_pendentes
    )
# COLABORADORES CADASTRADOS

@main_bp.route('/colaboradores-cadastrados')
def colaboradores_cadastrados():
    nucleo_filtro = request.args.get('nucleo', '').strip()

    with db_cursor() as cursor:
        if 'administrador_id' in session:
            cursor.execute("""
                SELECT DISTINCT nucleo
                FROM colaboradores
                WHERE trim(COALESCE(nucleo, '')) != ''
                ORDER BY nucleo
            """)
            nucleos = [row[0] for row in cursor.fetchall()]

            if nucleo_filtro:
                cursor.execute("""
                    SELECT *
                    FROM colaboradores
                    WHERE nucleo = ?
                    ORDER BY nome
                """, (nucleo_filtro,))
            else:
                cursor.execute("""
                    SELECT *
                    FROM colaboradores
                    ORDER BY nome
                """)

            colaboradores = cursor.fetchall()
            return render_template(
                'administrador/colaboradores.html',
                colaboradores=colaboradores,
                nucleos=nucleos,
                nucleo_filtro=nucleo_filtro
            )

        if 'gestor_id' in session:
            cursor.execute("""
                SELECT *
                FROM colaboradores
                WHERE nucleo = ?
                ORDER BY nome
            """, (session['nucleo_gestor'],))
            colaboradores = cursor.fetchall()
            return render_template(
                'gestor_nucleo/colaboradores.html',
                colaboradores=colaboradores
            )

    return redirect('/login-administrador')
# PONTO

@main_bp.route('/ponto')
def ponto():

    if 'colaborador_id' not in session:
        return redirect('/login-colaborador')

    data_hoje = _agora_sp().strftime('%d/%m/%Y')

    with db_cursor() as cursor:
        cursor.execute("""
            SELECT
                id,
                entrada,
                saida_final
            FROM registros_ponto
            WHERE colaborador_id = ?
            AND data = ?
        """, (
            session['colaborador_id'],
            data_hoje
        ))
        registro = cursor.fetchone()

        cursor.execute("""
            SELECT nucleo
            FROM colaboradores
            WHERE id = ?
        """, (session['colaborador_id'],))
        colaborador = cursor.fetchone()

        bloqueio_hoje = _bloqueio_trabalho(
            cursor,
            data_hoje,
            colaborador[0] if colaborador else ''
        )

    entrada = "--:--"
    saida_final = "--:--"
    entrada_registrada = False
    saida_registrada = False

    if registro:
        _, entrada_salva, saida_salva = registro
        entrada = entrada_salva or "--:--"
        saida_final = saida_salva or "--:--"
        entrada_registrada = bool((entrada_salva or '').strip())
        saida_registrada = bool((saida_salva or '').strip())

    return render_template(
        'colaborador/ponto.html',
        nome=session['nome'],
        entrada=entrada,
        saida_final=saida_final,
        entrada_registrada=entrada_registrada,
        saida_registrada=saida_registrada,
        erro=_mensagem_query('erro'),
        sucesso=_mensagem_query('sucesso'),
        bloqueio_hoje=bloqueio_hoje
    )
# AJUSTE DE PONTO

@main_bp.route('/ajuste-ponto')
def ajuste_ponto():

    if 'colaborador_id' not in session:
        return redirect('/login-colaborador')

    with db_cursor() as cursor:
        cursor.execute("""
            SELECT nucleo
            FROM colaboradores
            WHERE id = ?
        """, (session['colaborador_id'],))
        colaborador = cursor.fetchone()
        nucleo = colaborador[0] if colaborador else ''
        data_atual = _proximo_dia_util(cursor, nucleo).strftime('%Y-%m-%d')
        datas_feriado = _datas_feriado_nucleo(cursor, nucleo)
    
    return render_template(
        'colaborador/ajuste_ponto.html',
        data_atual=data_atual,
        datas_feriado=datas_feriado,
        erro=_mensagem_query('erro'),
        sucesso=_mensagem_query('sucesso')
    )
# SOLICITAÇÕES DE AJUSTE

@main_bp.route('/solicitacoes-ajuste')
def solicitacoes_ajuste():

    if 'administrador_id' not in session and 'gestor_id' not in session:
        return redirect('/login-gestor-nucleo')

    with db_cursor() as cursor:
        # Administrador vê todos
        if 'gestor_id' not in session:

            cursor.execute("""
                SELECT
                    ajustes_ponto.id,
                    colaboradores.nome,
                    colaboradores.nucleo,
                    ajustes_ponto.data,
                    ajustes_ponto.tipo_ajuste,
                    ajustes_ponto.horario_correto,
                    ajustes_ponto.motivo,
                    ajustes_ponto.status,
                    ajustes_ponto.motivo_reprovacao,
                    ajustes_ponto.analisado_por
                FROM ajustes_ponto
                INNER JOIN colaboradores
                ON ajustes_ponto.colaborador_id = colaboradores.id
                ORDER BY ajustes_ponto.id DESC
            """)

        # Gestor do Núcleo vê apenas os ajustes do seu núcleo
        else:

            cursor.execute("""
                SELECT
                    ajustes_ponto.id,
                    colaboradores.nome,
                    colaboradores.nucleo,
                    ajustes_ponto.data,
                    ajustes_ponto.tipo_ajuste,
                    ajustes_ponto.horario_correto,
                    ajustes_ponto.motivo,
                    ajustes_ponto.status,
                    ajustes_ponto.motivo_reprovacao,
                    ajustes_ponto.analisado_por
                FROM ajustes_ponto
                INNER JOIN colaboradores
                ON ajustes_ponto.colaborador_id = colaboradores.id
                WHERE colaboradores.nucleo = ?
                ORDER BY ajustes_ponto.id DESC
            """, (
                session['nucleo_gestor'],
            ))

        ajustes = cursor.fetchall()
    erro=_mensagem_query('erro')
    if 'administrador_id' in session:
        template_name = 'administrador/solicitacoes_ajuste.html'
    elif 'gestor_id' in session:
        template_name = 'gestor_nucleo/solicitacoes_ajuste.html'
    else:
        return redirect('/login-gestor-nucleo')

    return render_template(
        template_name,
        ajustes=ajustes,
        erro=erro
    )
# APROVAR AJUSTE

@main_bp.route('/aprovar-ajuste/<int:id>')
def aprovar_ajuste(id):

    with db_cursor(commit=True) as cursor:
        permissao = True
        nome_colaborador = 'Colaborador'

        # Segurança para Gestor do Núcleo
        if 'gestor_id' in session:

            cursor.execute("""
                SELECT ajustes_ponto.id
                FROM ajustes_ponto
                INNER JOIN colaboradores
                ON ajustes_ponto.colaborador_id = colaboradores.id
                WHERE ajustes_ponto.id = ?
                AND colaboradores.nucleo = ?
            """, (
                id,
                session['nucleo_gestor']
            ))

            permissao = cursor.fetchone()

        if not permissao:
            return redirect('/solicitacoes-ajuste')

        cursor.execute("""
            SELECT
                colaboradores.nome,
                ajustes_ponto.colaborador_id,
                ajustes_ponto.data,
                ajustes_ponto.tipo_ajuste,
                ajustes_ponto.horario_correto
            FROM ajustes_ponto
            INNER JOIN colaboradores ON colaboradores.id = ajustes_ponto.colaborador_id
            WHERE ajustes_ponto.id = ?
        """, (id,))
        ajuste_info = cursor.fetchone()

        if ajuste_info:
            nome_colaborador = ajuste_info[0]
            colaborador_id = ajuste_info[1]
            data_ajuste = ajuste_info[2]
            tipo_ajuste = (ajuste_info[3] or '').strip().lower()
            horario_correto = (ajuste_info[4] or '').strip()

            cursor.execute("""
                SELECT id, entrada, saida_final
                FROM registros_ponto
                WHERE colaborador_id = ?
                AND data = ?
            """, (
                colaborador_id,
                data_ajuste
            ))
            registro = cursor.fetchone()

            if registro:
                registro_id, entrada_atual, saida_atual = registro
                nova_entrada = entrada_atual
                nova_saida = saida_atual

                if tipo_ajuste == 'entrada':
                    nova_entrada = horario_correto
                elif tipo_ajuste in ('saida', 'saída'):
                    nova_saida = horario_correto

                cursor.execute("""
                    UPDATE registros_ponto
                    SET entrada = ?, saida_final = ?
                    WHERE id = ?
                """, (
                    nova_entrada,
                    nova_saida,
                    registro_id
                ))
            else:
                entrada = horario_correto if tipo_ajuste == 'entrada' else ''
                saida_final = horario_correto if tipo_ajuste in ('saida', 'saída') else ''

                cursor.execute("""
                    INSERT INTO registros_ponto (
                        colaborador_id,
                        data,
                        entrada,
                        saida_final
                    )
                    VALUES (?, ?, ?, ?)
                """, (
                    colaborador_id,
                    data_ajuste,
                    entrada,
                    saida_final
                ))

        cursor.execute("""
            UPDATE ajustes_ponto
            SET
                status = 'Aprovado',
                motivo_reprovacao = NULL,
                analisado_por = ?,
                data_analise = ?
            WHERE id = ?
        """, (
            _nome_reprovador(),
            _agora_sp().strftime('%d/%m/%Y %H:%M:%S'),
            id
        ))
        _registrar_log(
            cursor,
            'Aprovação de ajuste',
            f"{session.get('nome_admin') or session.get('nome_gestor', 'Usuário')} aprovou ajuste de {nome_colaborador}."
        )

    return redirect('/solicitacoes-ajuste')
# REPROVAR AJUSTE

@main_bp.route('/reprovar-ajuste/<int:id>', methods=['GET', 'POST'])
def reprovar_ajuste(id):

    motivo_reprovacao = request.form.get('motivo_reprovacao', '').strip()
    if not motivo_reprovacao:
        motivo_reprovacao = request.args.get('motivo', '').strip()
    if not motivo_reprovacao:
        motivo_reprovacao = 'Reprovado sem observação informada.'

    with db_cursor(commit=True) as cursor:
        nome_colaborador = 'Colaborador'

        # Segurança para Gestor do Núcleo
        if 'gestor_id' in session:
            cursor.execute("""
                SELECT ajustes_ponto.id
                FROM ajustes_ponto
                INNER JOIN colaboradores
                ON ajustes_ponto.colaborador_id = colaboradores.id
                WHERE ajustes_ponto.id = ?
                AND colaboradores.nucleo = ?
            """, (
                id,
                session['nucleo_gestor']
            ))

            permissao = cursor.fetchone()

            if not permissao:
                return redirect('/solicitacoes-ajuste')

        cursor.execute("""
            SELECT colaboradores.nome
            FROM ajustes_ponto
            INNER JOIN colaboradores ON colaboradores.id = ajustes_ponto.colaborador_id
            WHERE ajustes_ponto.id = ?
        """, (id,))
        ajuste_info = cursor.fetchone()
        if ajuste_info:
            nome_colaborador = ajuste_info[0]

        cursor.execute("""
            UPDATE ajustes_ponto
            SET
                status = 'Reprovado',
                motivo_reprovacao = ?,
                analisado_por = ?,
                data_analise = ?
            WHERE id = ?
        """, (
            motivo_reprovacao,
            _nome_reprovador(),
            _agora_sp().strftime('%d/%m/%Y %H:%M:%S'),
            id
        ))
        _registrar_log(
            cursor,
            'Reprovação de ajuste',
            f"{session.get('nome_admin') or session.get('nome_gestor', 'Usuário')} reprovou ajuste de {nome_colaborador}."
        )

    return redirect('/solicitacoes-ajuste')
# MEUS AJUSTES

@main_bp.route('/meus-ajustes')
def meus_ajustes():

    if 'colaborador_id' not in session:
        return redirect('/login-colaborador')

    with db_cursor() as cursor:
        cursor.execute("""
        SELECT
            data,
            tipo_ajuste,
            horario_correto,
            motivo,
            status,
            motivo_reprovacao,
            analisado_por,
            data_analise
        FROM ajustes_ponto
        WHERE colaborador_id = ?
        ORDER BY id DESC
    """, (
        session['colaborador_id'],
    ))

    ajustes = cursor.fetchall()
    return render_template(
        'colaborador/meus_ajustes.html',
        ajustes=ajustes
    )
# CALENDÁRIO

@main_bp.route('/calendario')
def calendario():

    # Data atual no formato YYYY-MM-DD para o input type="date"
    data_atual = datetime.now().strftime('%Y-%m-%d')

    with db_cursor() as cursor:
        if 'gestor_id' in session:
            cursor.execute("""
                SELECT id, titulo, tipo, data, descricao
                FROM eventos
                WHERE nucleo = ?
                ORDER BY data
            """, (
                session['nucleo_gestor'],
            ))
            eventos = cursor.fetchall()
            return render_template('gestor_nucleo/calendario.html', eventos=eventos, nucleo=session['nucleo_gestor'], data_atual=data_atual)

        if 'administrador_id' in session:
            cursor.execute("""
                SELECT id, titulo, tipo, data, descricao, nucleo
                FROM eventos
                ORDER BY data
            """)
        eventos = cursor.fetchall()
        return render_template('administrador/calendario.html', eventos=eventos, data_atual=data_atual)
    return redirect('/login-gestor-nucleo')


@main_bp.route('/adicionar-evento', methods=['POST'])
def adicionar_evento():

    if 'gestor_id' not in session:
        return redirect('/login-gestor-nucleo')

    titulo = request.form.get('titulo', '').strip()
    tipo = request.form.get('tipo', '').strip()
    data = request.form.get('data', '').strip()
    descricao = request.form.get('descricao', '').strip()

    if not titulo or not tipo or not data:
        return redirect('/calendario')

    try:
        data_formatada = datetime.strptime(data, '%Y-%m-%d').strftime('%d/%m/%Y')
    except ValueError:
        data_formatada = data

    with db_cursor(commit=True) as cursor:
        cursor.execute("""
        INSERT INTO eventos (
            titulo,
            tipo,
            data,
            descricao,
            nucleo
        ) VALUES (?, ?, ?, ?, ?)
    """, (
        titulo,
        tipo,
        data_formatada,
        descricao,
        session['nucleo_gestor']
    ))

    return redirect('/calendario')


@main_bp.route('/apagar-evento', methods=['POST'])
def apagar_evento():

    if 'gestor_id' not in session:
        return redirect('/login-gestor-nucleo')

    evento_id = request.form.get('event_id')
    if not evento_id:
        return redirect('/calendario')

    with db_cursor(commit=True) as cursor:
        cursor.execute("""
        DELETE FROM eventos
        WHERE id = ?
        AND nucleo = ?
    """, (
        evento_id,
        session['nucleo_gestor']
    ))

    return redirect('/calendario')


def _buscar_status_folhas(cursor, nucleo=None):
    if nucleo is None:
        cursor.execute("""
            SELECT id, nome, folha_assinada_path, folha_assinada_nome
            FROM colaboradores
            WHERE status = 'Ativo'
            ORDER BY nome
        """)
    else:
        cursor.execute("""
            SELECT id, nome, folha_assinada_path, folha_assinada_nome
            FROM colaboradores
            WHERE status = 'Ativo'
            AND nucleo = ?
            ORDER BY nome
        """, (nucleo,))

    return [
        {
            'id': row[0],
            'nome': row[1],
            'anexado': bool((row[2] or '').strip()),
            'arquivo_path': row[2],
            'arquivo_nome': row[3]
        }
        for row in cursor.fetchall()
    ]

def _todos_folhas_anexadas(status_folhas):
    """Verifica se todos os colaboradores têm folha anexada"""
    if not status_folhas:
        return False
    return all(item.get('anexado', False) for item in status_folhas)
# RELATÓRIOS

@main_bp.route('/relatorios')
def relatorios():

    with db_cursor() as cursor:
        # Administrador vê todos
        if 'gestor_id' not in session:

            cursor.execute("""
                SELECT id, nome
                FROM colaboradores
                WHERE status = 'Ativo'
                ORDER BY nome
            """)

        # Gestor do Núcleo vê apenas o seu núcleo
        else:

            cursor.execute("""
                SELECT id, nome
                FROM colaboradores
                WHERE status = 'Ativo'
                AND nucleo = ?
                ORDER BY nome
            """, (
                session['nucleo_gestor'],
            ))

        colaboradores = cursor.fetchall()
        status_folhas = _buscar_status_folhas(
            cursor,
            session.get('nucleo_gestor') if 'gestor_id' in session else None
        )
    if 'administrador_id' in session:
        template_name = 'administrador/relatorios.html'
    elif 'gestor_id' in session:
        template_name = 'gestor_nucleo/relatorios.html'
    else:
        return redirect('/login-gestor-nucleo')

    # Data atual no formato YYYY-MM para o input type="month"
    mes_atual = datetime.now().strftime('%Y-%m')

    return render_template(
        template_name,
        colaboradores=colaboradores,
        registros=None,
        todos=False,
        status_folhas=status_folhas,
        todos_anexados=_todos_folhas_anexadas(status_folhas),
        mes_atual=mes_atual,
        erro=_mensagem_query('erro')
    )


@main_bp.route('/autenticar-logs', methods=['POST'])
def autenticar_logs():

    if 'administrador_id' not in session:
        return redirect('/login-administrador')

    login = request.form.get('login_logs', '').strip()
    senha = request.form.get('senha_logs', '')

    if login == 'Pge.nati' and senha == 'Nati@ti2026':
        session['logs_auth_ok'] = True
        return redirect('/logs')

    return redirect('/dashboard-administrador?erro_logs=1')


@main_bp.route('/logs')
def logs_sistema():

    if 'administrador_id' not in session:
        return redirect('/login-administrador')

    if not session.get('logs_auth_ok'):
        return redirect('/dashboard-administrador?erro_logs=1')

    acao_filtro = request.args.get('acao', '').strip()
    texto_filtro = request.args.get('texto', '').strip()

    with db_cursor() as cursor:
        filtros = []
        parametros = []

        if acao_filtro:
            filtros.append("acao = ?")
            parametros.append(acao_filtro)

        if texto_filtro:
            termo = f"%{texto_filtro.lower()}%"
            filtros.append("(lower(nome_usuario) LIKE ? OR lower(acao) LIKE ? OR lower(detalhes) LIKE ?)")
            parametros.extend([termo, termo, termo])

        where_sql = ''
        if filtros:
            where_sql = 'WHERE ' + ' AND '.join(filtros)

        cursor.execute(f"""
            SELECT
                data_hora,
                tipo_usuario,
                nome_usuario,
                acao,
                detalhes
            FROM logs_sistema
            {where_sql}
            ORDER BY id DESC
            LIMIT 500
        """, parametros)
        logs = cursor.fetchall()

        cursor.execute("""
            SELECT DISTINCT acao
            FROM logs_sistema
            WHERE trim(COALESCE(acao, '')) != ''
            ORDER BY acao
        """)
        acoes = [row[0] for row in cursor.fetchall()]

    return render_template(
        'administrador/logs.html',
        logs=logs,
        acoes=acoes,
        acao_filtro=acao_filtro,
        texto_filtro=texto_filtro
    )
# VISUALIZAR RELATÓRIO

@main_bp.route('/visualizar-relatorio', methods=['POST'])
def visualizar_relatorio():

    colaborador_id = request.form['colaborador_id']

    mes = request.form.get('mes')

    with db_cursor() as cursor:
        # Lista de colaboradores
        if 'gestor_id' not in session:
            cursor.execute("""
                SELECT id, nome
                FROM colaboradores
                WHERE status = 'Ativo'
                ORDER BY nome
            """)
        else:
            cursor.execute("""
                SELECT id, nome
                FROM colaboradores
                WHERE status = 'Ativo'
                AND nucleo = ?
                ORDER BY nome
            """, (session['nucleo_gestor'],))

        colaboradores = cursor.fetchall()
        status_folhas = _buscar_status_folhas(
            cursor,
            session.get('nucleo_gestor') if 'gestor_id' in session else None
        )

        # Segurança adicional para Gestor do Núcleo
        if 'gestor_id' in session and colaborador_id != 'all':
            cursor.execute("""
                SELECT id
                FROM colaboradores
                WHERE id = ?
                AND nucleo = ?
            """, (
                colaborador_id,
                session['nucleo_gestor']
            ))
            permissao = cursor.fetchone()

            if not permissao:
                return render_template(
                    'gestor_nucleo/relatorios.html',
                    colaboradores=colaboradores,
                    registros=None,
                    erro='Você não possui permissão para visualizar este colaborador.',
                    mes=mes,
                    todos=False,
                    status_folhas=status_folhas,
                    todos_anexados=_todos_folhas_anexadas(status_folhas)
                )

        # Busca os registros — filtra por mês/ano se fornecido
        if mes:
            try:
                ano, mes_num = mes.split('-')
            except ValueError:
                ano = None
                mes_num = None
        else:
            ano = None
            mes_num = None

        if colaborador_id == 'all':
            if 'administrador_id' in session:
                filtro_todos = ""
                parametros = []
            elif 'gestor_id' in session:
                filtro_todos = " AND colaboradores.nucleo = ?"
                parametros = [session['nucleo_gestor']]
            else:
                return redirect('/login-administrador')

            if ano and mes_num:
                cursor.execute(f"""
                    SELECT
                        registros_ponto.data,
                        registros_ponto.entrada,
                        registros_ponto.saida_final,
                        COALESCE(NULLIF(registros_ponto.observacao, ''), group_concat(eventos.tipo || ' - ' || eventos.titulo, '; ')) AS observacao,
                        colaboradores.nome
                    FROM registros_ponto
                    INNER JOIN colaboradores ON colaboradores.id = registros_ponto.colaborador_id
                    LEFT JOIN eventos ON eventos.data = registros_ponto.data
                        AND eventos.nucleo = colaboradores.nucleo
                    WHERE substr(registros_ponto.data,4,2) = ?
                    AND substr(registros_ponto.data,7,4) = ?
                    {filtro_todos}
                    GROUP BY registros_ponto.id
                    ORDER BY registros_ponto.id DESC
                """, [mes_num, ano] + parametros)
            else:
                cursor.execute(f"""
                    SELECT
                        registros_ponto.data,
                        registros_ponto.entrada,
                        registros_ponto.saida_final,
                        COALESCE(NULLIF(registros_ponto.observacao, ''), group_concat(eventos.tipo || ' - ' || eventos.titulo, '; ')) AS observacao,
                        colaboradores.nome
                    FROM registros_ponto
                    INNER JOIN colaboradores ON colaboradores.id = registros_ponto.colaborador_id
                    LEFT JOIN eventos ON eventos.data = registros_ponto.data
                        AND eventos.nucleo = colaboradores.nucleo
                    WHERE 1 = 1
                    {filtro_todos}
                    GROUP BY registros_ponto.id
                    ORDER BY registros_ponto.id DESC
                """, parametros)

            registros = cursor.fetchall()
            nome_colaborador = 'Todos os colaboradores'
            todos = True
        else:
            registros = _registros_folha_colaborador(
                cursor,
                colaborador_id,
                mes_num if (ano and mes_num) else None,
                ano if (ano and mes_num) else None,
                ordem='desc'
            )

            cursor.execute("""
                SELECT nome
                FROM colaboradores
                WHERE id = ?
            """, (colaborador_id,))
            nome_colaborador_row = cursor.fetchone()
            nome_colaborador = nome_colaborador_row[0] if nome_colaborador_row else 'Colaborador'
            todos = False
    template_name = 'administrador/relatorios.html'
    if 'gestor_id' in session:
        template_name = 'gestor_nucleo/relatorios.html'

    return render_template(
        template_name,
        colaboradores=colaboradores,
        registros=registros,
        nome_colaborador=nome_colaborador,
        mes=mes,
        todos=todos,
        colaborador_id=colaborador_id,
        status_folhas=status_folhas,
        todos_anexados=_todos_folhas_anexadas(status_folhas)
    )
# SALVAR COLABORADOR

@main_bp.route('/salvar-colaborador', methods=['POST'])
def salvar_colaborador():

    with db_cursor(commit=True) as cursor:
        # Se for Gestor do Núcleo, usa automaticamente o núcleo dele
        if 'gestor_id' in session:
            nucleo = session['nucleo_gestor']
        # Se for Administrador, permite escolher o núcleo
        else:
            nucleo = request.form['nucleo']

        cursor.execute("""
            INSERT INTO colaboradores (
                nome,
                vinculo,
                nucleo,
                turno,
                horario,
                presencial,
                unidade_exercicio,
                procurador_monitor,
                celular,
                login,
                senha
            )
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
    """, (
        request.form['nome'],
        request.form['vinculo'],
        nucleo,
        request.form['turno'],
        request.form['horario'],
        request.form['presencial'],
        request.form['unidade_exercicio'],
        request.form['procurador_monitor'],
        request.form['celular'],
        request.form['login'],
        request.form['senha']
    ))
        _registrar_log(
            cursor,
            'Cadastro de colaborador',
            f"{session.get('nome_admin', 'Usuário')} cadastrou o colaborador {request.form['nome']}."
        )

    return redirect('/colaboradores-cadastrados')
# CANCELAR COLABORADOR

@main_bp.route('/cancelar-colaborador/<int:id>', methods=['POST'])
def cancelar_colaborador(id):

    motivo = request.form.get('motivo_cancelamento', '').strip()

    with db_cursor(commit=True) as cursor:
        cursor.execute("SELECT nome FROM colaboradores WHERE id = ?", (id,))
        colaborador_info = cursor.fetchone()
        cursor.execute("""
        UPDATE colaboradores
        SET status = 'Inativo', cancel_observacao = ?
        WHERE id = ?
    """, (
        motivo,
        id,
    ))
        _registrar_log(
            cursor,
            'Cancelamento de colaborador',
            f"{session.get('nome_admin', 'Administrador')} cancelou o colaborador {(colaborador_info[0] if colaborador_info else id)}."
        )

    return redirect('/colaboradores-cadastrados')


@main_bp.route('/recadastrar-colaborador/<int:id>', methods=['POST'])
def recadastrar_colaborador(id):

    with db_cursor(commit=True) as cursor:
        cursor.execute("SELECT nome FROM colaboradores WHERE id = ?", (id,))
        colaborador_info = cursor.fetchone()
        cursor.execute("""
        UPDATE colaboradores
        SET status = 'Ativo', cancel_observacao = NULL
        WHERE id = ?
    """, (
        id,
    ))
        _registrar_log(
            cursor,
            'Recadastro de colaborador',
            f"{session.get('nome_admin', 'Administrador')} reativou o colaborador {(colaborador_info[0] if colaborador_info else id)}."
        )

    return redirect('/colaboradores-cadastrados')
# EDITAR COLABORADOR

@main_bp.route('/editar-colaborador/<int:id>')
def editar_colaborador(id):

    with db_cursor() as cursor:
        cursor.execute("""
        SELECT *
        FROM colaboradores
        WHERE id = ?
    """, (
        id,
    ))

    colaborador = cursor.fetchone()
    return render_template(
        'administrador/editar_colaborador.html',
        colaborador=colaborador
    )
# ATUALIZAR COLABORADOR

@main_bp.route('/atualizar-colaborador/<int:id>', methods=['POST'])
def atualizar_colaborador(id):

    with db_cursor(commit=True) as cursor:
        cursor.execute("""
        UPDATE colaboradores
        SET
            nome = ?,
            nucleo = ?,
            turno = ?,
            horario = ?,
            celular = ?
        WHERE id = ?
    """, (
        request.form['nome'],
        request.form['nucleo'],
        request.form['turno'],
        request.form['horario'],
        request.form['celular'],
        id
    ))

    return redirect('/colaboradores-cadastrados')
# REGISTRAR ENTRADA

@main_bp.route('/registrar-entrada', methods=['POST'])
def registrar_entrada():

    if 'colaborador_id' not in session:
        return redirect('/login-colaborador')

    with db_cursor(commit=True) as cursor:
        hora_atual = request.form['entrada']
        data_hoje = _agora_sp().strftime('%d/%m/%Y')

        cursor.execute("""
            SELECT nucleo
            FROM colaboradores
            WHERE id = ?
        """, (session['colaborador_id'],))
        colaborador = cursor.fetchone()
        bloqueio = _bloqueio_trabalho(
            cursor,
            data_hoje,
            colaborador[0] if colaborador else ''
        )
        if bloqueio:
            return _redirect_com_mensagem('/ponto', 'erro', bloqueio)

        cursor.execute("""
            SELECT id, entrada, saida_final
            FROM registros_ponto
            WHERE colaborador_id = ?
            AND data = ?
        """, (session['colaborador_id'], data_hoje))

        registro = cursor.fetchone()

        if not registro:
            cursor.execute("""
                INSERT INTO registros_ponto (
                    colaborador_id,
                    data,
                    entrada
                )
                VALUES (?, ?, ?)
            """, (
                session['colaborador_id'],
                data_hoje,
                hora_atual
            ))
            _registrar_log(
                cursor,
                'Registro de entrada',
                f"{session.get('nome', 'Colaborador')} registrou entrada em {data_hoje} às {hora_atual}."
            )
        else:
            registro_id, entrada_salva, saida_salva = registro
            if not (entrada_salva or '').strip():
                cursor.execute("""
                    UPDATE registros_ponto
                    SET entrada = ?
                    WHERE id = ?
                """, (
                    hora_atual,
                    registro_id
                ))
                _registrar_log(
                    cursor,
                    'Atualização de entrada',
                    f"{session.get('nome', 'Colaborador')} atualizou entrada em {data_hoje} para {hora_atual}."
                )
    return _redirect_com_mensagem('/ponto', 'sucesso', 'Entrada registrada com sucesso.')
# REGISTRAR SAÍDA FINAL

@main_bp.route('/registrar-saida-final', methods=['POST'])
def registrar_saida_final():

    if 'colaborador_id' not in session:
        return redirect('/login-colaborador')

    with db_cursor(commit=True) as cursor:
        hora_atual = request.form['saida_final']
        data_hoje = _agora_sp().strftime('%d/%m/%Y')

        cursor.execute("""
            SELECT nucleo
            FROM colaboradores
            WHERE id = ?
        """, (session['colaborador_id'],))
        colaborador = cursor.fetchone()
        bloqueio = _bloqueio_trabalho(
            cursor,
            data_hoje,
            colaborador[0] if colaborador else ''
        )
        if bloqueio:
            return _redirect_com_mensagem('/ponto', 'erro', bloqueio)

        cursor.execute("""
            SELECT id, entrada, saida_final
            FROM registros_ponto
            WHERE colaborador_id = ?
            AND data = ?
        """, (
            session['colaborador_id'],
            data_hoje
        ))

        registro = cursor.fetchone()

        if registro:
            registro_id, entrada_salva, saida_salva = registro
            if not (saida_salva or '').strip():
                cursor.execute("""
                    UPDATE registros_ponto
                    SET saida_final = ?
                    WHERE id = ?
                """, (
                    hora_atual,
                    registro_id
                ))
                _registrar_log(
                    cursor,
                    'Registro de saída',
                    f"{session.get('nome', 'Colaborador')} registrou saída em {data_hoje} às {hora_atual}."
                )
    return _redirect_com_mensagem('/ponto', 'sucesso', 'Saída registrada com sucesso.')

@main_bp.route('/meu-relatorio')
def meu_relatorio():

    if 'colaborador_id' not in session:
        return redirect('/login-colaborador')

    mes_atual = datetime.now(
        ZoneInfo("America/Sao_Paulo")
    ).strftime('%Y-%m')
    mes_selecionado = request.args.get('mes') or mes_atual
    ano = mes_selecionado.split('-')[0]
    mes = mes_selecionado.split('-')[1]

    with db_cursor() as cursor:
        registros = _registros_folha_colaborador(
            cursor,
            session['colaborador_id'],
            mes,
            ano,
            ordem='desc'
        )

        cursor.execute("""
            SELECT folha_assinada_nome
            FROM colaboradores
            WHERE id = ?
        """, (session['colaborador_id'],))
        folha_assinada = cursor.fetchone()

    data_atual = datetime.now(
        ZoneInfo("America/Sao_Paulo")
    ).strftime('%Y-%m-%d')
    
    return render_template(
        'colaborador/meu_relatorio.html',
        registros=registros,
        nome=session['nome'],
        mes_selecionado=mes_selecionado,
        data_atual=data_atual,
        folha_assinada_nome=folha_assinada[0] if folha_assinada else None,
        mes_atual=mes_atual
    )


@main_bp.route('/anexar-folha-assinada', methods=['POST'])
def anexar_folha_assinada():

    if 'colaborador_id' not in session:
        return redirect('/login-colaborador')

    arquivo = request.files.get('arquivo_folha')
    if not arquivo or arquivo.filename == '':
        return redirect('/meu-relatorio')

    if not arquivo.filename.lower().endswith('.pdf'):
        return redirect('/meu-relatorio')

    pasta_upload = ensure_upload_folder()

    nome_seguro = safe_filename(arquivo.filename)
    nome_arquivo = f"{session['colaborador_id']}_{nome_seguro}"
    caminho_arquivo = os.path.join(pasta_upload, nome_arquivo)
    arquivo.save(caminho_arquivo)

    with db_cursor(commit=True) as cursor:
        cursor.execute("""
        UPDATE colaboradores
        SET folha_assinada_path = ?, folha_assinada_nome = ?
        WHERE id = ?
    """, (
        caminho_arquivo,
        nome_seguro,
        session['colaborador_id']
    ))
    return redirect('/meu-relatorio')


@main_bp.route('/download-folhas-assinadas')
def download_folhas_assinadas():

    if 'administrador_id' not in session and 'gestor_id' not in session:
        return redirect('/login-gestor-nucleo')

    with db_cursor() as cursor:
        if 'gestor_id' in session:
            cursor.execute("""
                SELECT id, nome, folha_assinada_path
                FROM colaboradores
                WHERE status = 'Ativo'
                AND nucleo = ?
                AND folha_assinada_path IS NOT NULL
                AND trim(folha_assinada_path) != ''
                ORDER BY nome
            """, (session['nucleo_gestor'],))
        else:
            cursor.execute("""
                SELECT id, nome, folha_assinada_path
                FROM colaboradores
                WHERE status = 'Ativo'
                AND folha_assinada_path IS NOT NULL
                AND trim(folha_assinada_path) != ''
                ORDER BY nome
            """)

        anexos = cursor.fetchall()
    if not anexos:
        return redirect('/relatorios')

    pasta_upload = ensure_upload_folder()

    nome_zip = os.path.join(pasta_upload, f"pacote_folhas_assinadas_{datetime.now().strftime('%Y%m%d_%H%M%S')}.zip")

    with ZipFile(nome_zip, 'w') as zip_file:
        for colaborador_id, nome, caminho in anexos:
            if os.path.exists(caminho):
                nome_arquivo_zip = f"{nome.replace('/', '_').replace('\\', '_')}_{colaborador_id}.pdf"
                zip_file.write(caminho, arcname=nome_arquivo_zip)

    return send_file(
        nome_zip,
        as_attachment=True,
        download_name='pacote_folhas_assinadas.zip'
    )


@main_bp.route('/download-folha-assinada/<int:colaborador_id>')
def download_folha_assinada_individual(colaborador_id):

    if 'administrador_id' not in session and 'gestor_id' not in session:
        return redirect('/login-gestor-nucleo')

    with db_cursor() as cursor:
        if 'gestor_id' in session:
            cursor.execute("""
                SELECT nome, folha_assinada_path, folha_assinada_nome
                FROM colaboradores
                WHERE id = ?
                AND nucleo = ?
            """, (
                colaborador_id,
                session['nucleo_gestor']
            ))
        else:
            cursor.execute("""
                SELECT nome, folha_assinada_path, folha_assinada_nome
                FROM colaboradores
                WHERE id = ?
            """, (colaborador_id,))

        colaborador = cursor.fetchone()

    if not colaborador:
        return _redirect_com_mensagem('/relatorios', 'erro', 'Colaborador não encontrado.')

    nome, caminho, nome_arquivo = colaborador
    if not (caminho and str(caminho).strip()) or not os.path.exists(caminho):
        return _redirect_com_mensagem('/relatorios', 'erro', f'Folha assinada não encontrada para {nome}.')

    download_name = nome_arquivo or f"folha_assinada_{colaborador_id}.pdf"
    return send_file(
        caminho,
        as_attachment=True,
        download_name=download_name,
        mimetype='application/pdf'
    )


@main_bp.route('/salvar-observacao', methods=['POST'])
def salvar_observacao():

    if 'colaborador_id' not in session:
        return redirect('/login-colaborador')

    data = request.form['data']
    observacao = request.form['observacao'].strip()
    data_convertida = datetime.strptime(
        data,
        "%Y-%m-%d"
    ).strftime("%d/%m/%Y")

    with db_cursor(commit=True) as cursor:
        cursor.execute("""
            UPDATE registros_ponto
            SET observacao = ?
            WHERE colaborador_id = ?
            AND data = ?
        """, (
            observacao,
            session['colaborador_id'],
            data_convertida
        ))
        _registrar_log(
            cursor,
            'Observação de ponto',
            f"{session.get('nome', 'Colaborador')} definiu observação '{observacao}' para {data_convertida}."
        )

    return redirect('/meu-relatorio')

@main_bp.route('/gerar-pdf')
def gerar_pdf():

    colaborador_id = request.args.get('colaborador_id')
    mes_selecionado = request.args.get('mes')

    if 'administrador_id' not in session and 'gestor_id' not in session and 'colaborador_id' not in session:
        return redirect('/login-colaborador')

    with db_cursor() as cursor:
        if 'administrador_id' in session or 'gestor_id' in session:
            if not colaborador_id:
                return redirect('/relatorios')

            if colaborador_id == 'all':
                colaborador = ('Todos os Colaboradores', '', '')
            else:
                cursor.execute("""
                    SELECT
                        nome,
                        nucleo,
                        horario
                    FROM colaboradores
                    WHERE id = ?
                """, (colaborador_id,))
                colaborador = cursor.fetchone()

                if not colaborador:
                    return redirect('/relatorios')

                if 'gestor_id' in session:
                    cursor.execute("""
                        SELECT id
                        FROM colaboradores
                        WHERE id = ?
                        AND nucleo = ?
                    """, (colaborador_id, session['nucleo_gestor']))
                    if not cursor.fetchone():
                        return redirect('/relatorios')

        else:
            colaborador_id = session['colaborador_id']
            cursor.execute("""
                SELECT
                    nome,
                    nucleo,
                    horario
                FROM colaboradores
                WHERE id = ?
            """, (colaborador_id,))
            colaborador = cursor.fetchone()

    print("MES PDF:", mes_selecionado)

    if colaborador_id == 'all':
        filtro_nucleo = ""
        parametros_extra = []
        if 'gestor_id' in session:
            filtro_nucleo = " AND colaboradores.nucleo = ?"
            parametros_extra.append(session['nucleo_gestor'])

        if mes_selecionado and "-" in mes_selecionado:
            ano, mes = mes_selecionado.split('-')
            cursor.execute(f"""
                SELECT
                    registros_ponto.data,
                    registros_ponto.entrada,
                    registros_ponto.saida_final,
                    COALESCE(NULLIF(registros_ponto.observacao, ''), group_concat(eventos.tipo || ' - ' || eventos.titulo, '; ')) AS observacao,
                    colaboradores.nome
                FROM registros_ponto
                INNER JOIN colaboradores ON colaboradores.id = registros_ponto.colaborador_id
                LEFT JOIN eventos ON eventos.data = registros_ponto.data
                    AND eventos.nucleo = colaboradores.nucleo
                WHERE substr(registros_ponto.data,4,2) = ?
                AND substr(registros_ponto.data,7,4) = ?
                {filtro_nucleo}
                GROUP BY registros_ponto.id
                ORDER BY registros_ponto.data
            """, [mes, ano] + parametros_extra)
        else:
            cursor.execute(f"""
                SELECT
                    registros_ponto.data,
                    registros_ponto.entrada,
                    registros_ponto.saida_final,
                    COALESCE(NULLIF(registros_ponto.observacao, ''), group_concat(eventos.tipo || ' - ' || eventos.titulo, '; ')) AS observacao,
                    colaboradores.nome
                FROM registros_ponto
                INNER JOIN colaboradores ON colaboradores.id = registros_ponto.colaborador_id
                LEFT JOIN eventos ON eventos.data = registros_ponto.data
                    AND eventos.nucleo = colaboradores.nucleo
                WHERE 1 = 1
                {filtro_nucleo}
                GROUP BY registros_ponto.id
                ORDER BY registros_ponto.data
            """, parametros_extra)
        registros = cursor.fetchall()
        nome_colaborador = colaborador[0].replace(" ", "_")

        if mes_selecionado and "-" in mes_selecionado:
            ano = mes_selecionado.split('-')[0]
            mes = mes_selecionado.split('-')[1]
            arquivo = f"Folha_Frequencia_Todos_Colaboradores_{mes}_{ano}.pdf"
        else:
            arquivo = f"Folha_Frequencia_Todos_Colaboradores.pdf"
    else:
        if mes_selecionado and "-" in mes_selecionado:
            ano, mes = mes_selecionado.split('-')
            registros = _registros_folha_colaborador(
                cursor,
                colaborador_id,
                mes,
                ano,
                ordem='asc'
            )
        else:
            registros = _registros_folha_colaborador(
                cursor,
                colaborador_id,
                ordem='asc'
            )
        nome_colaborador = colaborador[0].replace(" ", "_")

        if mes_selecionado and "-" in mes_selecionado:
            ano = mes_selecionado.split('-')[0]
            mes = mes_selecionado.split('-')[1]
            arquivo = f"Folha_Frequencia_{nome_colaborador}_{mes}_{ano}.pdf"
        else:
            arquivo = f"Folha_Frequencia_{nome_colaborador}.pdf"

    if colaborador_id == 'all':
        temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.pdf')
        temp_path = temp_file.name
        temp_file.close()
        doc = SimpleDocTemplate(temp_path)
        estilos = getSampleStyleSheet()
        elementos = []
        elementos.append(
            Paragraph(
                "<b>FOLHA DE FREQUÊNCIA</b>",
                estilos['Title']
            )
        )
        if mes_selecionado and "-" in mes_selecionado:
            elementos.append(
                Paragraph(
                    f"<b>Competência:</b> {mes}/{ano}",
                    estilos['Normal']
                )
            )
            elementos.append(Spacer(1, 12))
        elementos.append(
            Paragraph(
                f"<b>Relatório:</b> Todos os colaboradores",
                estilos['Normal']
            )
        )
        elementos.append(Spacer(1, 20))
        estilo_celula_tabela = ParagraphStyle(
            name='CelulaTabelaPDF',
            parent=estilos['Normal'],
            fontSize=8,
            leading=10,
            wordWrap='CJK'
        )
        estilo_cabecalho_tabela = ParagraphStyle(
            name='CabecalhoTabelaPDF',
            parent=estilos['Normal'],
            fontName='Helvetica-Bold',
            fontSize=8,
            leading=10,
            alignment=1,
            wordWrap='CJK'
        )

        dados = [
            [
                _celula_pdf("Data", estilo_cabecalho_tabela),
                _celula_pdf("Entrada", estilo_cabecalho_tabela),
                _celula_pdf("Saída", estilo_cabecalho_tabela),
                _celula_pdf("Observação", estilo_cabecalho_tabela),
                _celula_pdf("Colaborador", estilo_cabecalho_tabela)
            ]
        ]
        for registro in registros:
            dados.append([
                _celula_pdf(registro[0], estilo_celula_tabela),
                _celula_pdf(registro[1] or "", estilo_celula_tabela),
                _celula_pdf(registro[2] or "", estilo_celula_tabela),
                _celula_pdf(registro[3] or "", estilo_celula_tabela),
                _celula_pdf(registro[4] or "", estilo_celula_tabela)
            ])
        tabela = Table(
            dados,
            colWidths=[64, 50, 50, 230, 126],
            repeatRows=1
        )
        tabela.setStyle(TableStyle([
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ]))
        elementos.append(tabela)
        doc.build(elementos)
    else:
        temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.pdf')
        temp_path = temp_file.name
        temp_file.close()
        doc = SimpleDocTemplate(
            temp_path,
            pagesize=A4,
            leftMargin=24,
            rightMargin=24,
            topMargin=20,
            bottomMargin=20
        )
        estilos = getSampleStyleSheet()
        elementos = []
        elementos.append(
            Paragraph(
                "<b>FOLHA DE FREQUÊNCIA</b>",
                estilos['Title']
            )
        )
        if mes_selecionado and "-" in mes_selecionado:
            elementos.append(
                Paragraph(
                    f"<b>Competência:</b> {mes}/{ano}",
                    estilos['Normal']
                )
            )
            elementos.append(Spacer(1, 12))
        elementos.append(
            Paragraph(
                f"<b>Nome:</b> {colaborador[0]}",
                estilos['Normal']
            )
        )
        elementos.append(
            Paragraph(
                f"<b>Núcleo:</b> {colaborador[1]}",
                estilos['Normal']
            )
        )
        elementos.append(
            Paragraph(
                f"<b>Horário:</b> {colaborador[2]}",
                estilos['Normal']
            )
        )
        elementos.append(Spacer(1, 12))
        estilo_celula_tabela = ParagraphStyle(
            name='CelulaTabelaPDF',
            parent=estilos['Normal'],
            fontSize=8,
            leading=10,
            wordWrap='CJK'
        )
        estilo_cabecalho_tabela = ParagraphStyle(
            name='CabecalhoTabelaPDF',
            parent=estilos['Normal'],
            fontName='Helvetica-Bold',
            fontSize=8,
            leading=10,
            alignment=1,
            wordWrap='CJK'
        )

        dados = [
            [
                _celula_pdf("Data", estilo_cabecalho_tabela),
                _celula_pdf("Entrada", estilo_cabecalho_tabela),
                _celula_pdf("Saída", estilo_cabecalho_tabela),
                _celula_pdf("Observação", estilo_cabecalho_tabela)
            ]
        ]
        for registro in registros:
            dados.append([
                _celula_pdf(registro[0], estilo_celula_tabela),
                _celula_pdf(registro[1] or "", estilo_celula_tabela),
                _celula_pdf(registro[2] or "", estilo_celula_tabela),
                _celula_pdf(registro[3] or "", estilo_celula_tabela)
            ])
        tabela = Table(
            dados,
            colWidths=[78, 58, 58, 120],
            hAlign='CENTER',
            repeatRows=1
        )
        tabela.setStyle(TableStyle([
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('LEADING', (0, 0), (-1, -1), 10),
            ('LEFTPADDING', (0, 0), (-1, -1), 4),
            ('RIGHTPADDING', (0, 0), (-1, -1), 4),
            ('TOPPADDING', (0, 0), (-1, -1), 3),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ]))
        elementos.append(tabela)
        elementos.append(Spacer(1, 20))
        elementos.append(
            Paragraph(
                "<b>VISTO</b>",
                estilos['Title']
            )
        )
        elementos.append(Spacer(1, 12))
        assinaturas = Table([
            [
                "__________________________________",
                "__________________________________"
            ],
            [
                colaborador[0],
                "ASSINATURA/CARIMBO"
            ],
            [
                "Relatório Abrangente",
                "PROCURADOR DO ESTADO"
            ]
        ], colWidths=[260, 260])
        assinaturas.setStyle(TableStyle([
            ('ALIGN', (0,0), (-1,-1), 'CENTER'),
            ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
            ('FONTNAME', (0,0), (-1,-1), 'Helvetica'),
            ('FONTSIZE', (0,0), (-1,-1), 10),
            ('BOTTOMPADDING', (0,0), (-1,-1), 6),
        ]))
        elementos.append(assinaturas)
        doc.build(elementos)

    clean_temp_file(temp_path)

    return send_file(
        temp_path,
        as_attachment=True,
        download_name=arquivo,
        mimetype='application/pdf'
    )

@main_bp.route('/gerar-excel')
def gerar_excel():

    colaborador_id = request.args.get('colaborador_id')
    mes_selecionado = request.args.get('mes')

    if 'administrador_id' not in session and 'gestor_id' not in session and 'colaborador_id' not in session:
        return redirect('/login-colaborador')

    with db_cursor() as cursor:
        if 'administrador_id' in session or 'gestor_id' in session:
            if not colaborador_id:
                return redirect('/relatorios')

            if colaborador_id == 'all':
                colaborador = ('Todos os Colaboradores', '', '')
            else:
                cursor.execute("""
                    SELECT
                        nome,
                        nucleo,
                        horario
                    FROM colaboradores
                    WHERE id = ?
                """, (colaborador_id,))
                colaborador = cursor.fetchone()
                if not colaborador:
                    return redirect('/relatorios')
                if 'gestor_id' in session:
                    cursor.execute("""
                        SELECT id
                        FROM colaboradores
                        WHERE id = ?
                        AND nucleo = ?
                    """, (colaborador_id, session['nucleo_gestor']))
                    if not cursor.fetchone():
                        return redirect('/relatorios')
        else:
            colaborador_id = session['colaborador_id']
            cursor.execute("""
                SELECT
                    nome,
                    nucleo,
                    horario
                FROM colaboradores
                WHERE id = ?
            """, (colaborador_id,))
            colaborador = cursor.fetchone()

    print("MES EXCEL:", mes_selecionado)

    if colaborador_id == 'all':
        filtro_nucleo = ""
        parametros_extra = []
        if 'gestor_id' in session:
            filtro_nucleo = " AND colaboradores.nucleo = ?"
            parametros_extra.append(session['nucleo_gestor'])

        if mes_selecionado and "-" in mes_selecionado:
            ano, mes = mes_selecionado.split('-')
            cursor.execute(f"""
                SELECT
                    data,
                    entrada,
                    saida_final,
                    observacao,
                    colaboradores.nome
                FROM registros_ponto
                INNER JOIN colaboradores ON colaboradores.id = registros_ponto.colaborador_id
                WHERE substr(registros_ponto.data,4,2) = ?
                AND substr(registros_ponto.data,7,4) = ?
                {filtro_nucleo}
                ORDER BY data
            """, [mes, ano] + parametros_extra)
        else:
            cursor.execute(f"""
                SELECT
                    data,
                    entrada,
                    saida_final,
                    observacao,
                    colaboradores.nome
                FROM registros_ponto
                INNER JOIN colaboradores ON colaboradores.id = registros_ponto.colaborador_id
                WHERE 1 = 1
                {filtro_nucleo}
                ORDER BY data
            """, parametros_extra)
    else:
        if mes_selecionado and "-" in mes_selecionado:
            ano, mes = mes_selecionado.split('-')
            registros = _registros_folha_colaborador(
                cursor,
                colaborador_id,
                mes,
                ano,
                ordem='asc'
            )
        else:
            registros = _registros_folha_colaborador(
                cursor,
                colaborador_id,
                ordem='asc'
            )

    if colaborador_id == 'all':
        registros = cursor.fetchall()
    wb = Workbook()
    ws = wb.active
    ws.title = "Frequência"
    ws['A1'] = "FOLHA DE FREQUÊNCIA"

    if colaborador_id == 'all':
        if mes_selecionado and "-" in mes_selecionado:
            ano = mes_selecionado.split('-')[0]
            mes = mes_selecionado.split('-')[1]
            ws['A2'] = f"Competência: {mes}/{ano}"
        ws['A3'] = "Relatório"
        ws['B3'] = "Todos os Colaboradores"
        ws['A4'] = "Núcleo"
        ws['B4'] = "Todos"
        ws['A5'] = "Horário"
        ws['B5'] = "Todos"
        ws.append([])
        ws.append([
            "Data",
            "Entrada",
            "Saída",
            "Observação",
            "Colaborador"
        ])
        for registro in registros:
            ws.append([
                registro[0],
                registro[1] or "",
                registro[2] or "",
                registro[3] or "",
                registro[4] or ""
            ])
        if mes_selecionado and "-" in mes_selecionado:
            arquivo = f"Folha_Frequencia_Todos_Colaboradores_{mes}_{ano}.xlsx"
        else:
            arquivo = f"Folha_Frequencia_Todos_Colaboradores.xlsx"
    else:
        if mes_selecionado and "-" in mes_selecionado:
            ano = mes_selecionado.split('-')[0]
            mes = mes_selecionado.split('-')[1]
            ws['A2'] = f"Competência: {mes}/{ano}"
        ws['A3'] = "Nome"
        ws['B3'] = colaborador[0]
        ws['A4'] = "Núcleo"
        ws['B4'] = colaborador[1]
        ws['A5'] = "Horário"
        ws['B5'] = colaborador[2]
        ws.append([])
        ws.append([
            "Data",
            "Entrada",
            "Saída",
            "Observação"
        ])
        for registro in registros:
            ws.append([
                registro[0],
                registro[1] or "",
                registro[2] or "",
                registro[3] or ""
            ])
        nome_colaborador = colaborador[0].replace(" ", "_")
        if mes_selecionado and "-" in mes_selecionado:
            arquivo = f"Folha_Frequencia_{nome_colaborador}_{mes}_{ano}.xlsx"
        else:
            arquivo = f"Folha_Frequencia_{nome_colaborador}.xlsx"

    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(
                wrap_text=True,
                vertical='top'
            )

    temp_path = save_temp_file('.xlsx')
    wb.save(temp_path)

    clean_temp_file(temp_path)

    return send_file(
        temp_path,
        as_attachment=True,
        download_name=arquivo,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

@main_bp.route('/primeiro-acesso')
def primeiro_acesso():

    if 'colaborador_id' not in session:
        return redirect('/login-colaborador')

    return render_template(
    'colaborador/primeiro_acesso.html'
)

@main_bp.route('/alterar-senha', methods=['POST'])
def alterar_senha():

    if 'colaborador_id' not in session:
        return redirect('/login-colaborador')

    senha1 = request.form['senha']
    senha2 = request.form['confirmar_senha']

    if senha1 != senha2:
        return """
        As senhas não coincidem.
        """

    if senha1 == "Novocolab123":
        return """
        A senha padrão não pode ser utilizada novamente.
        """

    padrao = r'^(?=.*[a-z])(?=.*[A-Z])(?=.*\d)(?=.*[@$!%*?&]).{8,}$'

    if not re.match(padrao, senha1):

        return """
        A senha deve possuir:

            - mínimo de 8 caracteres;
        - uma letra maiúscula;
        - uma letra minúscula;
        - um número;
        - um caractere especial.
        """

    with db_cursor(commit=True) as cursor:
        cursor.execute("""
        UPDATE colaboradores
        SET senha = ?
        WHERE id = ?
    """, (
        senha1,
        session['colaborador_id']
    ))

    return redirect('/ponto')

@main_bp.route('/lancar-ponto')
def lancar_ponto():

    if 'administrador_id' not in session:
        return redirect('/login-administrador')

    with db_cursor() as cursor:
        cursor.execute("""
        SELECT DISTINCT trim(nucleo) AS nucleo
        FROM colaboradores
        WHERE status = 'Ativo'
        AND trim(COALESCE(nucleo, '')) != ''
        ORDER BY nucleo
    """)
        nucleos = [row[0] for row in cursor.fetchall()]

        cursor.execute("""
        SELECT id, nome, trim(COALESCE(nucleo, ''))
        FROM colaboradores
        WHERE status = 'Ativo'
        ORDER BY nome
    """)
        colaboradores = cursor.fetchall()
    
    # Data atual no formato YYYY-MM-DD para o input type="date"
    data_atual = datetime.now().strftime('%Y-%m-%d')
    
    return render_template(
        'administrador/lancar_ponto.html',
        colaboradores=colaboradores,
        nucleos=nucleos,
        data_atual=data_atual,
        erro=_mensagem_query('erro'),
        sucesso=_mensagem_query('sucesso')
    )

@main_bp.route('/salvar-lancamento-ponto', methods=['POST'])
def salvar_lancamento_ponto():

    if 'administrador_id' not in session:
        return redirect('/login-administrador')

    colaborador_id = request.form['colaborador_id']
    data = request.form['data']
    entrada = request.form['entrada']
    saida_final = request.form['saida_final']
    observacao = request.form['observacao']
    data_formatada = _data_iso_para_br(data)

    with db_cursor(commit=True) as cursor:
        cursor.execute("SELECT nome, nucleo FROM colaboradores WHERE id = ?", (colaborador_id,))
        colaborador_info = cursor.fetchone()
        bloqueio = _bloqueio_trabalho(
            cursor,
            data_formatada,
            colaborador_info[1] if colaborador_info else ''
        )
        if bloqueio:
            return _redirect_com_mensagem('/lancar-ponto', 'erro', bloqueio)

        cursor.execute("""
            SELECT id
            FROM registros_ponto
            WHERE colaborador_id = ?
            AND data = ?
        """, (
            colaborador_id,
            data_formatada
        ))
        registro_existente = cursor.fetchone()

        if registro_existente:
            cursor.execute("""
                UPDATE registros_ponto
                SET
                    entrada = ?,
                    saida_final = ?,
                    observacao = ?
                WHERE id = ?
            """, (
                entrada,
                saida_final,
                observacao,
                registro_existente[0]
            ))
        else:
            cursor.execute("""
            INSERT INTO registros_ponto (
                colaborador_id,
                data,
                entrada,
                saida_final,
                observacao
            )
            VALUES (?, ?, ?, ?, ?)
        """, (
            colaborador_id,
            data_formatada,
            entrada,
            saida_final,
            observacao
        ))
        _registrar_log(
            cursor,
            'Lançamento manual de ponto',
            f"{session.get('nome_admin', 'Administrador')} lançou ponto para {(colaborador_info[0] if colaborador_info else colaborador_id)} em {data_formatada}."
        )

    return _redirect_com_mensagem('/lancar-ponto', 'sucesso', 'Ponto lançado com sucesso.')

@main_bp.route('/enviar-ajuste', methods=['POST'])
def enviar_ajuste():

    if 'colaborador_id' not in session:
        return redirect('/login-colaborador')

    data_ajuste = _data_iso_para_br(request.form.get('data'))
    tipo_ajuste = request.form.get('tipo', '').strip()
    horario_correto = request.form.get('horario', '').strip()
    motivo = request.form.get('motivo', '').strip()

    if tipo_ajuste not in ('Entrada', 'Saída'):
        return _redirect_com_mensagem('/ajuste-ponto', 'erro', 'Tipo de ajuste inválido.')

    with db_cursor(commit=True) as cursor:
        cursor.execute("""
            SELECT nucleo
            FROM colaboradores
            WHERE id = ?
        """, (session['colaborador_id'],))
        colaborador = cursor.fetchone()

        bloqueio = _bloqueio_trabalho(
            cursor,
            data_ajuste,
            colaborador[0] if colaborador else ''
        )
        if bloqueio:
            return _redirect_com_mensagem('/ajuste-ponto', 'erro', bloqueio)

        cursor.execute("""
        INSERT INTO ajustes_ponto (
            colaborador_id,
            data,
            tipo_ajuste,
            horario_correto,
            motivo,
            status
        )
        VALUES (?, ?, ?, ?, ?, ?)
    """, (
        session['colaborador_id'],
        data_ajuste,
        tipo_ajuste,
        horario_correto,
        motivo,
        'Pendente'
    ))
        _registrar_log(
            cursor,
            'Solicitação de ajuste',
            f"{session.get('nome', 'Colaborador')} abriu solicitação de ajuste para {data_ajuste}."
        )

    return _redirect_com_mensagem('/ajuste-ponto', 'sucesso', 'Solicitação enviada com sucesso.')

@main_bp.route('/primeiro-acesso-gestor')
def primeiro_acesso_gestor():

    if 'gestor_id' not in session:
        return redirect('/login-gestor-nucleo')

    return render_template(
        'gestor_nucleo/primeiro_acesso_gestor.html'
    )

@main_bp.route('/alterar-senha-gestor', methods=['POST'])
def alterar_senha_gestor():

    if 'gestor_id' not in session:
        return redirect('/login-gestor-nucleo')

    senha1 = request.form['senha']
    senha2 = request.form['confirmar_senha']

    if senha1 != senha2:
        return "As senhas não coincidem."

    if senha1 == "Novocolab123":
        return "A senha padrão não pode ser utilizada novamente."

    padrao = r'^(?=.*[a-z])(?=.*[A-Z])(?=.*\d)(?=.*[@$!%*?&]).{8,}$'

    if not re.match(padrao, senha1):

        return """
        A senha deve possuir:

            - mínimo de 8 caracteres;
        - uma letra maiúscula;
        - uma letra minúscula;
        - um número;
        - um caractere especial.
        """

    with db_cursor(commit=True) as cursor:
        cursor.execute("""
        UPDATE gestores
        SET senha = ?
        WHERE id = ?
    """, (
        senha1,
        session['gestor_id']
    ))

    return redirect('/dashboard-gestor-nucleo')
